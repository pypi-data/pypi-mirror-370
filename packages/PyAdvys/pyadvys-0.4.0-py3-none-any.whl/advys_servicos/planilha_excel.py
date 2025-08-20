import logging
import openpyxl
from pathlib import Path

import pandas as pd
from advys_validadores.arquivo import verifica_arquivo


class ArquivoExcel:
    # Construtor e inicizizção
    def __init__(self, caminho: Path):
        self.caminho_arquivo = caminho
        self.nome_arquivo = caminho.name
        self.cria_arquivo_xlsx()

    def cria_arquivo_xlsx(self) -> None:
        extensao_valida = self.caminho_arquivo.suffix.lower() == ".xlsx"
        if self.caminho_arquivo.exists() and extensao_valida:
            self.arquivo = openpyxl.load_workbook(self.caminho_arquivo)
            logging.info(
                f"Arquivo {self.caminho_arquivo.name} carregado."
            )
        else:
            self.arquivo = openpyxl.Workbook()
            logging.info(f"Arquivo {self.caminho_arquivo.name} criado.")

    # Métodos principais
    def cria_planilha(self, nome_planilha: str) -> None:
        self._seleciona_ou_cria_planilha(nome_planilha)

    def selecionar_planilha(self, nome_planilha: str) -> None:
        self._seleciona_ou_cria_planilha(nome_planilha)

    def adiciona_nome_coluna(self, nome: str, coluna: int) -> None:
        if not isinstance(coluna, int) or coluna < 1:
            raise ValueError(f"Coluna {coluna} inválida!")
        self.planilha.cell(row=1, column=coluna, value=nome)

    def adiciona_linha(self, linha: list) -> None:
        if linha:
            self.planilha.append(linha)
        else:
            raise ValueError(f"Lista vazia")

    def adiciona_dataframe(self, df: pd.DataFrame) -> None:
        # Insere nome das colunas
        for col_num, value in enumerate(df.columns, start=1):
            self.adiciona_nome_coluna(value, col_num)

        # Insere valores nas linhas
        for linha in df.itertuples(index=False):
            self.planilha.append(linha)

    def salva_arquivo(self) -> None:
        if "Sheet" in self.arquivo.sheetnames:
            self._remover_sheet_padrao()
        self.arquivo.save(str(self.caminho_arquivo))

    # Métodos utilitários
    def limpa_planilha(self, manter_cabecalho: bool = False) -> None:
        if manter_cabecalho and self.planilha.max_row > 1:
            self.planilha.delete_rows(2, self.planilha.max_row - 1)
        else:
            self.planilha.delete_rows(1, self.planilha.max_row)

    def formata_tamanho_colunas(self) -> None:
        for coluna in range(1, self.planilha.max_column + 1):
            max_length = 10
            for linha in range(1, self.planilha.max_row + 1):
                cell = self.planilha.cell(row=linha, column=coluna)
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            self.planilha.column_dimensions[
                self.planilha.cell(row=1, column=coluna).column_letter
            ].width = max_length + 2

    def remove_planilha(self, nome_planilha: str = None) -> None:
        try:
            planilha = (
                self.arquivo[nome_planilha]
                if nome_planilha else self.planilha
            )
            self.arquivo.remove(planilha)
        except Exception as e:
            logging.error(
                f"Erro ao excluir planilha '{planilha.title}': {e}"
            )

    # Métodos Privados
    def _seleciona_ou_cria_planilha(self, nome_planilha: str) -> None:
        try:
            # Verifica se planilha já existe
            if nome_planilha in self.arquivo.sheetnames:
                self.planilha = self.arquivo[nome_planilha]
            else:
                self.planilha = self.arquivo.create_sheet(nome_planilha)
                logging.info(f"Planilha {nome_planilha} criada.")
        except Exception as e:
            logging.error(f"Erro ao criar planilha {nome_planilha}: {e}")

    def _remover_sheet_padrao(self) -> None:
        try:
            planilha_padrao = self.arquivo["Sheet"]
            self.arquivo.remove(planilha_padrao)
        except KeyError:
            pass  # Sheet já foi removida

    # Gerenciador de contexto
    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        try:
            self.salva_arquivo()
        except Exception as e:
            logging.error(f"Erro ao salvar arquivo: {e}")
