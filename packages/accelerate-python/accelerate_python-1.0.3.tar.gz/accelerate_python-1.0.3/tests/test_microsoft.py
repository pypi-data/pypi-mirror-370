## external imports
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

import pytest
from assertpy import assert_that
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill

## internal imports
from accelerate.python import CellFormat, ExcelUtil, WrapperTest

## global variables
INPUT_PATH = Path(__file__).parent.joinpath("input")


## test imports
class TestCellFormat(WrapperTest):
    @pytest.fixture
    def cell_format(self) -> CellFormat:
        return CellFormat(
            font=Font(name="Arial", size=12, bold=True),
            fill=PatternFill(patternType="solid", fgColor="FF0000"),
            border=Border(),
            number_format="0.00",
        )

    def test_default_initialization(self):
        cell_format = CellFormat()
        assert_that(cell_format.font).is_none()
        assert_that(cell_format.fill).is_none()
        assert_that(cell_format.border).is_none()
        assert_that(cell_format.number_format).is_none()
        assert_that(cell_format.protection).is_none()
        assert_that(cell_format.alignment).is_none()
        assert_that(cell_format.style).is_none()

    def test_custom_initialization(self, cell_format: CellFormat):
        assert_that(cell_format.font.name).is_equal_to("Arial")
        assert_that(cell_format.font.size).is_equal_to(12)
        assert_that(cell_format.font.bold).is_true()
        assert_that(cell_format.fill.patternType).is_equal_to("solid")
        assert_that(cell_format.fill.fgColor.rgb).is_equal_to("00FF0000")
        assert_that(cell_format.number_format).is_equal_to("0.00")

    def test_update(self, cell_format: CellFormat):
        cell_format.font = Font(name="Calibri", size=11)
        assert_that(cell_format.font.name).is_equal_to("Calibri")
        assert_that(cell_format.font.size).is_equal_to(11)

        cell_format.fill = PatternFill(patternType="solid", fgColor="00FF00")
        assert_that(cell_format.fill.patternType).is_equal_to("solid")
        assert_that(cell_format.fill.fgColor.rgb).is_equal_to("0000FF00")


class TestExcelUtil(WrapperTest):
    @pytest.fixture
    def workbook(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["headA", "headB", "headC"])
        worksheet.append(["dataA", "dataB", "dataC"])
        return workbook

    def test_new_workbook_default(self):
        # Test creating a new workbook with default arguments
        wb = ExcelUtil.new_workbook()
        assert_that(wb).is_instance_of(Workbook)
        assert_that(wb.worksheets, "worksheets").is_length(
            1
        )  # Default workbook has 1 sheet

    def test_new_workbook_custom_options(self):
        # Test creating a new workbook with custom options
        wb = ExcelUtil.new_workbook(write_only=True, iso_dates=True)
        assert_that(wb).is_instance_of(Workbook)
        assert_that(wb.write_only, "write_only").is_true()
        assert_that(wb.iso_dates, "iso_dates").is_true()

    def test_load_workbook(self):
        # read a test workbook
        test_workbook = ExcelUtil.load_workbook(INPUT_PATH.joinpath("test.xlsx"))
        assert_that(test_workbook).is_instance_of(Workbook)
        assert_that(test_workbook.worksheets).is_not_empty()

    @patch("openpyxl.load_workbook", return_value=Workbook())
    def test_load_workbook_options(self, mock_load_workbook):
        # Mock load_workbook with custom options
        wb = ExcelUtil.load_workbook("fake_path.xlsx", True, True, True, False, True)
        mock_load_workbook.assert_called_once_with(
            "fake_path.xlsx",
            read_only=True,
            keep_vba=True,
            data_only=True,
            keep_links=False,
            rich_text=True,
        )
        assert_that(wb).is_instance_of(Workbook)

    @patch("openpyxl.workbook.workbook.Workbook.save")
    def test_save_workbook_to_file(self, mock_workbook_save):
        # Mock saving a workbook
        path = Path("fake_path.xlsx")
        output = ExcelUtil.write_workbook(Workbook(), path.as_posix())
        mock_workbook_save.assert_called_once_with(path)
        assert_that(output, "Output File").is_equal_to(path)

    def test_save_workbook_to_bytes(self):
        # Mock saving a workbook
        content = (
            assert_that(ExcelUtil.write_workbook_bytes(Workbook()), "Workbook Bytes")
            .is_instance_of(BytesIO)
            .val
        )
        assert_that(content.getvalue()).is_not_empty()

    def test_format_row(self, workbook: Workbook):
        worksheet = workbook.active
        worksheet.append(["dataA", "dataB", "dataC"])
        row = worksheet[worksheet.max_row]
        ExcelUtil.format_row(
            row,
            CellFormat(
                font=Font(bold=True, color="ffffff"),
                fill=PatternFill(
                    start_color="cb6015", end_color="cb6015", fill_type="solid"
                ),
            ),
            {3: CellFormat(number_format="mm/dd/yy;@")},
        )
        for cell in row:
            assert_that(cell.font.bold).is_true()
            assert_that(cell.fill.start_color.index).is_equal_to("00cb6015")
        assert_that(row[2].number_format).is_equal_to("mm/dd/yy;@")

    def test_add_table(self):
        worksheet = Workbook().active
        worksheet.append(["headA", "headB", "headC"])
        worksheet.append(["dataA", "dataB", "dataC"])
        ExcelUtil.add_table(
            worksheet,
            "Table1",
            "A1:C2",
            "TableStyleMedium2",
            showRowStripes=True,
        )
        assert_that(worksheet.tables).contains_key("Table1")
        table = worksheet.tables["Table1"]
        assert_that(table, "New Table").has_displayName("Table1").has_ref("A1:C2")
        assert_that(table.tableStyleInfo.showRowStripes).is_true()

    def test_resize_table(self):
        worksheet = Workbook().active
        worksheet.append(["headA", "headB", "headC"])
        worksheet.append(["dataA", "dataB", "dataC"])
        ExcelUtil.add_table(
            worksheet,
            "Table1",
            "A1:B2",
            "TableStyleMedium2",
            showRowStripes=True,
        )
        ExcelUtil.resize_table(worksheet, "Table1")
        table = worksheet.tables["Table1"]
        assert_that(table, "Resized Table").has_ref("A1:C2")

    def test_auto_adjust_column_width(self):
        worksheet = Workbook().active
        worksheet.append(["headA", "headB", "headC"])
        worksheet.append(["dataA", "dataB", "dataC"])
        ExcelUtil.auto_adjust_column_width(worksheet)
        for col in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            assert_that(
                worksheet.column_dimensions[col[0].column_letter].width
            ).is_equal_to(max_length + 1)

    def test_column_letter(self):
        assert_that(ExcelUtil.column_letter(1)).is_equal_to("A")
        assert_that(ExcelUtil.column_letter(26)).is_equal_to("Z")
        assert_that(ExcelUtil.column_letter(27)).is_equal_to("AA")
        assert_that(ExcelUtil.column_letter(52)).is_equal_to("AZ")
        assert_that(ExcelUtil.column_letter(53)).is_equal_to("BA")


if __name__ == "__main__":
    pytest.main([__file__])
