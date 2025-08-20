"""This module provides wrapper classes to work with Microsoft documents - Excel, Word, Powerpoint"""

## external imports
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import TypeAlias

import openpyxl
from openpyxl.cell import Cell, MergedCell
from openpyxl.styles import Border, Font, PatternFill, styleable
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

## internal imports
from ._dataclasses import DataClass
from ._logging import AccelerateLogger
from ._tracker import PerformanceTracker

## performance tracking
PerformanceTracker.register_module_start(__name__)


## global variables
_LOGGER = AccelerateLogger.get_logger(__name__)
_CellOrMergedCell: TypeAlias = Cell | MergedCell


## class definitions
@dataclass()
class CellFormat(DataClass):
    font: Font | None = None
    fill: PatternFill | None = None
    border: Border | None = None
    number_format: str | None = None
    protection: styleable.StyleDescriptor | None = None
    alignment: styleable.StyleDescriptor | None = None
    style: styleable.NamedStyleDescriptor | None = None
    quotePrefix: styleable.StyleArrayDescriptor | None = None
    pivotButton: styleable.StyleArrayDescriptor | None = None


@dataclass()
class WorkbookTemplate(DataClass):
    name: str
    columns: list
    rowFormat: CellFormat | None = None
    columnFormats: dict[int, CellFormat] = field(default_factory=dict)


@_LOGGER.audit_class()
class ExcelUtil:
    """Wrapper class to manage repo operations"""

    @staticmethod
    def new_workbook(write_only=False, iso_dates=False) -> Workbook:
        return Workbook(write_only=write_only, iso_dates=iso_dates)

    @staticmethod
    def load_workbook(
        source: str | Path | BytesIO,
        read_only=False,
        keep_vba=False,
        data_only=False,
        keep_links=True,
        rich_text=False,
    ) -> Workbook:
        """
        :param source: Excel file path or byte stream | Required
        :param remaining params are for load_workbook() and optional
        """
        return openpyxl.load_workbook(
            source,
            read_only=read_only,
            keep_vba=keep_vba,
            data_only=data_only,
            keep_links=keep_links,
            rich_text=rich_text,
        )

    @staticmethod
    def write_workbook_bytes(workbook: Workbook) -> BytesIO:
        target = BytesIO()
        workbook.save(target)
        return target

    @staticmethod
    def write_workbook(workbook: Workbook, path: str | Path) -> Path:
        target = Path(path)
        workbook.save(target)
        return target

    @staticmethod
    def format_row(
        row: tuple[_CellOrMergedCell, ...],
        row_format: CellFormat | None = None,
        column_formats: dict[int, CellFormat] = {},
    ):
        for cell in row:
            if row_format:
                for key, value in vars(row_format).items():
                    if value:
                        setattr(cell, key, value)
            column_format = column_formats.get(cell.column) if cell.column else None
            if column_format:
                for key, value in vars(column_format).items():
                    if value:
                        setattr(cell, key, value)

    @staticmethod
    def add_table(
        sheet: Worksheet,
        table_name: str,
        range: str,
        style: str,
        showFirstColumn=None,
        showLastColumn=None,
        showRowStripes=None,
        showColumnStripes=None,
    ):
        table = Table(displayName=table_name, ref=range)
        table.tableStyleInfo = TableStyleInfo(
            name=style,
            showFirstColumn=showFirstColumn,
            showLastColumn=showLastColumn,
            showRowStripes=showRowStripes,
            showColumnStripes=showColumnStripes,
        )
        sheet.add_table(table)

    @staticmethod
    def resize_table(sheet: Worksheet, table_name: str):
        last_row = sheet[sheet.max_row]
        assert isinstance(last_row[-1], Cell)
        sheet.tables[table_name].ref = f"A1:{last_row[-1].column_letter}{sheet.max_row}"

    @staticmethod
    def auto_adjust_column_width(sheet: Worksheet):
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column = list(column)  # Convert the tuple to a list
            for cell in column:
                try:
                    # Find the length of the longest cell content
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass

            assert isinstance(column[0], Cell)
            sheet.column_dimensions[column[0].column_letter].width = max_length + 1

    @staticmethod
    def column_letter(column_number: int) -> str:
        """Convert column number to letter"""
        result = ""
        while column_number > 0:
            column_number, remainder = divmod(column_number - 1, 26)
            result = chr(65 + remainder) + result
        return result

    ## known formats
    FORMAT_CURRENCY = CellFormat(
        number_format='_([$$-en-US]* #,##0.00_);_([$$-en-US]* (#,##0.00);_([$$-en-US]* "-"??_);_(@_)'
    )
    FORMAT_DATE = CellFormat(number_format="mm/dd/yy;@")
    FORMAT_PERCENTAGE = CellFormat(number_format="0.00%")


## export symbols
__all__ = ["CellFormat", "WorkbookTemplate", "ExcelUtil"]


## log performance tracker
PerformanceTracker.log_module_load(__name__)


## disable standalone execution for library
if __name__ == "__main__":
    raise Exception(
        "This is not a standalone module, and should be imported as a library"
    )
