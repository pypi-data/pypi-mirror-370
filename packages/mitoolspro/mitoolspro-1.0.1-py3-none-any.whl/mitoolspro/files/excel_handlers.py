from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


def auto_adjust_excel_columns_width(
    excel_path: Path, max_len: int | None = None
) -> None:
    book = openpyxl.load_workbook(excel_path)
    for sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        auto_adjust_sheet_columns_width(sheet, max_len)
    book.save(excel_path)


def auto_adjust_sheet_columns_width(
    sheet: Worksheet, max_len: int | None = None
) -> None:
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column if cell.value]  # Filter out None values
        if not column:  # Skip if column is empty
            continue
        for cell in column:
            try:
                cell_len = max([len(line) for line in str(cell.value).split("\n")])
                if cell_len > max_length and (max_len is None or cell_len < max_len):
                    max_length = cell_len
            except Exception:
                pass
        adjusted_width = max_length + 1  # Adding a little extra width
        sheet.column_dimensions[
            openpyxl.utils.get_column_letter(column[0].column)
        ].width = adjusted_width
