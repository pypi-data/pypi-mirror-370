import logging
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment


class ExcelReporter:
    def __init__(self) -> None:
        self.wb = Workbook()
        self.logger = logging.getLogger()
    
    def auto_width(self, worksheet, customized={}):
        dims = {}
        for row in worksheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))

        for col, value in dims.items():
            self.logger.debug("Sheet:%s , column: %s, width: %d" % (str(worksheet), col, value))
            if col in customized:
                if customized[col] != 0:
                    worksheet.column_dimensions[col].width = customized[col]
            else:
                worksheet.column_dimensions[col].width = value + 4

    def write_title_row(self, sheet: Worksheet, title_row):
        for idx in range(0, len(title_row)):
            cell = sheet.cell(row=1, column=idx + 1)
            cell.value = title_row[idx]
            cell.alignment = Alignment(horizontal="center")

    def write_cell(self, sheet: Worksheet, row: int, column: int, value, format=None) -> Cell:
        cell = sheet.cell(row=row, column=column)         # type: Cell
        cell.value = value
        if (format is not None):
            if ('alignment' in format):
                cell.alignment = format['alignment']
            if ('number_format' in format):
                cell.number_format = format['number_format']
        return cell
    
    def write_cell_center(self, sheet: Worksheet, row: int, column: int, value) -> Cell:
        self.write_cell(sheet, row, column, value, format={"alignment": Alignment(horizontal="center")})
    
    def write_bool_cell(self, sheet: Worksheet, row: int, column: int, value) -> Cell:
        self.write_cell_center(sheet, row, column, self.format_boolean(value))
    
    def format_boolean(self, value: bool) -> str:
        if value is True:
            return "Y"
        else:
            return ""

    def save(self, name: str):
        self.wb.save(name)
