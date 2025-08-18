from ...models.eb_doc import EBModel
from ...models.nvm_xdm import NvMEaRef, NvMFeeRef
from ...reporter.excel_reporter.abstract import ExcelReporter
from openpyxl.styles.alignment import Alignment


class NvMXdmXlsWriter(ExcelReporter):
    def __init__(self) -> None:
        super().__init__()

    def write_nvm_general(self, doc: EBModel):
        sheet = self.wb.create_sheet("General", 0)

        title_row = ["Key", "Value"]
        self.write_title_row(sheet, title_row)

        nvm_common = doc.getNvM().getNvMCommon()

        if nvm_common is None:
            self.logger.error("NvMCommon is Invalid and General updating is skipped.")
            return

        row = 2
        self.write_cell(sheet, row, 1, "NvMCompiledConfigId")
        self.write_cell_center(sheet, row, 2, nvm_common.getNvMCompiledConfigId())
        row += 1
        self.write_cell(sheet, row, 1, "NvMDatasetSelectionBits")
        self.write_cell_center(sheet, row, 2, nvm_common.getNvMDatasetSelectionBits())
        row += 1
        self.write_cell(sheet, row, 1, "NvMMaxNumOfReadRetries")
        self.write_cell_center(sheet, row, 2, "3")
        row += 1
        self.write_cell(sheet, row, 1, "NvMMaxNumOfWriteRetries")
        self.write_cell_center(sheet, row, 2, "3")
        row += 1

        self.auto_width(sheet)

    def write_nvm_block_descriptors(self, doc: EBModel):
        sheet = self.wb.create_sheet("Block List", 2)

        title_row = [
            "BlockId", "Name", "NvMBlockEcucPartitionRef", "NvMRamBlockDataAddress", "NvMRomBlockDataAddress",
            "NvMBlockJobPriority", "NvMResistantToChangedSw", "NvMBlockManagementType", "NvMNvBlockLength", "NvMBlockCrcType",
            "NvMNvBlockNum", "NvMSelectBlockForReadAll", "NvMSelectBlockForWriteAll", "NvMProvideRteJobFinishedPort", "NvMProvideRteServicePort",
            "NvMInitBlockCallbackFnc", "NvMSingleBlockCallbackFnc", "NvMReadRamBlockFromNvCallback", "NvMWriteRamBlockToNvCallback", "NvMBlockUseSyncMechanism",  # noqa E501
            "NvMNvBlockBaseNumber", "NvMFeeRef"]
        self.write_title_row(sheet, title_row)

        row = 2
        for nvm_block in doc.getNvM().getNvMBlockDescriptorList():
            self.write_cell_center(sheet, row, 1, nvm_block.getNvMNvramBlockIdentifier())
            self.write_cell(sheet, row, 2, nvm_block.getName())
            if nvm_block.getNvMBlockEcucPartitionRef() is not None:
                self.write_cell_center(sheet, row, 3, nvm_block.getNvMBlockEcucPartitionRef().getShortName())
            self.write_cell(sheet, row, 4, nvm_block.getNvMRamBlockDataAddress())
            self.write_cell(sheet, row, 5, nvm_block.getNvMRomBlockDataAddress())

            self.write_cell_center(sheet, row, 6, nvm_block.getNvMBlockJobPriority())
            self.write_bool_cell(sheet, row, 7, nvm_block.getNvMResistantToChangedSw())
            self.write_cell_center(sheet, row, 8, nvm_block.getNvMBlockManagementType())
            self.write_cell_center(sheet, row, 9, nvm_block.getNvMNvBlockLength())
            if nvm_block.getNvMBlockUseCrc():
                self.write_cell_center(sheet, row, 10, nvm_block.getNvMBlockCrcType())

            self.write_cell(sheet, row, 11, nvm_block.getNvMNvBlockNum())
            self.write_bool_cell(sheet, row, 12, nvm_block.getNvMSelectBlockForReadAll())
            self.write_bool_cell(sheet, row, 13, nvm_block.getNvMSelectBlockForWriteAll())
            self.write_bool_cell(sheet, row, 14, nvm_block.getNvMProvideRteJobFinishedPort())
            self.write_bool_cell(sheet, row, 15, nvm_block.getNvMProvideRteServicePort())

            if nvm_block.getNvMInitBlockCallback() is not None:
                self.write_cell(sheet, row, 16, nvm_block.getNvMInitBlockCallback().getNvMInitBlockCallbackFnc())
            if nvm_block.getNvMSingleBlockCallback() is not None:
                self.write_cell(sheet, row, 17, nvm_block.getNvMSingleBlockCallback().getNvMSingleBlockCallbackFnc())
            if nvm_block.getNvMReadRamBlockFromNvCallback() is not None:
                self.write_cell(sheet, row, 18, nvm_block.getNvMReadRamBlockFromNvCallback())
            if nvm_block.getNvMWriteRamBlockToNvCallback() is not None:
                self.write_cell(sheet, row, 19, nvm_block.getNvMWriteRamBlockToNvCallback())
            self.write_bool_cell(sheet, row, 20, nvm_block.getNvMBlockUseSyncMechanism())
            
            self.write_cell_center(sheet, row, 21, nvm_block.getNvMNvBlockBaseNumber())
            block_reference = nvm_block.getNvMTargetBlockReference()
            if block_reference is not None:
                if isinstance(block_reference, NvMFeeRef):
                    self.write_cell(sheet, row, 22, block_reference.getNvMNameOfFeeBlock().getShortName())
                else:
                    raise NotImplementedError("Unsupported Target block reference.")
            
            row += 1

            self.logger.debug("Write NvM Block <%s>" % nvm_block.getName())

        self.auto_width(sheet)

    def write_nvm_bsw_distribution(self, doc: EBModel):
        sheet = self.wb.create_sheet("BSW Distribution", 1)

        title_row = ["NvMEcucPartitionRef", "Master"]
        self.write_title_row(sheet, title_row)

        nvm_common = doc.getNvM().getNvMCommon()

        if nvm_common is None:
            self.logger.error("NvMCommon is Invalid and BSW Distribution updating is skipped.")
            return
        
        master_partition_ref = nvm_common.getNvMMasterEcucPartitionRef()

        row = 2
        for ref_link in nvm_common.getNvMEcucPartitionRefList():
            self.write_cell(sheet, row, 1, ref_link.getShortName())
            if ref_link.getShortName() == master_partition_ref.getShortName():
                self.write_cell(sheet, row, 2, "Y", format={"alignment": Alignment(horizontal="center")})
            else:
                self.write_cell(sheet, row, 2, "N", format={"alignment": Alignment(horizontal="center")})
            row += 1

            self.logger.debug("Write NvM EcucPartition <%s>" % ref_link.getShortName())

        self.auto_width(sheet)

    def write(self, filename, doc: EBModel, options):
        self.logger.info("Writing <%s>" % filename)

        self.write_nvm_general(doc)
        self.write_nvm_bsw_distribution(doc)
        self.write_nvm_block_descriptors(doc)

        self.save(filename)
