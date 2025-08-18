import re
from openpyxl.styles import Alignment
from ...models.eb_doc import EBModel
from ...reporter.excel_reporter.abstract import ExcelReporter


class BswMXdmXlsWriter(ExcelReporter):
    def __init__(self) -> None:
        super().__init__()

    def write_os_tasks(self, doc: EBModel):
        sheet = self.wb.create_sheet("OsTask", 0)

        title_row = [
            "Name", "OsApplication", "OsTaskActivation", "OsTaskPriority", "OsTaskAutostart",
            "OsTaskSchedule", "OsStacksize", "OsTaskType", "OsResourceRef"]
        self.write_title_row(sheet, title_row)

        row = 2
        for os_task in doc.getOs().getOsTaskList():
            self.write_cell(sheet, row, 1, os_task.getName())
            os_app = doc.getOs().getOsTaskOsApplication(os_task.getName())
            self.write_cell(sheet, row, 2, os_app.getName())
            self.write_cell(sheet, row, 3, os_task.getOsTaskActivation())
            self.write_cell(sheet, row, 4, os_task.getOsTaskPriority())
            self.write_cell(sheet, row, 5, os_task.isOsTaskAutostart())

            self.write_cell(sheet, row, 6, os_task.getOsTaskSchedule())
            self.write_cell(sheet, row, 7, os_task.getOsStacksize())
            self.write_cell(sheet, row, 8, os_task.getOsTaskType())
            resources = []
            for resource_ref in os_task.getOsTaskResourceRefList():
                m = re.match(r"Rte_\w+", resource_ref.getValue())
                if m:
                    resources.append(resource_ref.getValue())
            total_resources = len(resources)
            if total_resources > 10:
                cell = self.write_cell(sheet, row, 9, "Total: %d OsResources" % (total_resources))
            else:
                cell = self.write_cell(sheet, row, 9, "\n".join(resources))
            if total_resources > 1 and total_resources < 10:
                cell.alignment = Alignment(wrapText=True)
            row += 1

            self.logger.debug("Write OsTask <%s>" % os_task.getName())

        self.auto_width(sheet)

    def write_os_isrs(self, doc: EBModel):
        sheet = self.wb.create_sheet("OsIsr", 1)

        title_row = [
            "Name", "OsApplication", "OsIsrCategory", "OsStacksize", "OsIsrPriority",
            "OsIsrVector", "MkMemoryRegion"]
        self.write_title_row(sheet, title_row)

        row = 2
        for os_isr in doc.getOs().getOsIsrList():
            self.write_cell(sheet, row, 1, os_isr.getName(), {'alignment': Alignment(vertical="top")})
            os_app = doc.getOs().getOsIsrOsApplication(os_isr.getName())
            if os_app is not None:
                self.write_cell(sheet, row, 2, os_app.getName(),
                                format={'alignment': Alignment(horizontal="center", vertical="top")})
            self.write_cell(sheet, row, 3, os_isr.getOsIsrCategory(),
                            format={'alignment': Alignment(horizontal="center", vertical="top")})
            self.write_cell(sheet, row, 4, os_isr.getOsStacksize(),
                            format={'alignment': Alignment(horizontal="center", vertical="top")})
            self.write_cell(sheet, row, 5, os_isr.getOsIsrPriority(),
                            format={'alignment': Alignment(horizontal="center", vertical="top")})
            self.write_cell(sheet, row, 6, os_isr.getOsIsrVector(),
                            format={'alignment': Alignment(horizontal="center", vertical="top")})
            if len(os_isr.getOsIsrMkMemoryRegionRefs()) > 1:
                self.write_cell(sheet, row, 7, "\n".join(map(lambda a: a.getShortName(), os_isr.getOsIsrMkMemoryRegionRefs())),
                                {'alignment': Alignment(wrapText=True, vertical="top")})
            else:
                self.write_cell(sheet, row, 7, "\n".join(map(lambda a: a.getShortName(), os_isr.getOsIsrMkMemoryRegionRefs())),
                                {'alignment': Alignment(vertical="top")})
            row += 1

            self.logger.debug("Write OsIsr <%s>" % os_isr.getName())

        self.auto_width(sheet, {"G": 25})

    def write_os_schedule_tables(self, doc: EBModel):
        sheet = self.wb.create_sheet("OsScheduleTable", 2)
        
        title_row = ["Name", "Duration", "Repeating", "OsCount"]
        self.write_title_row(sheet, title_row)

        row = 2
        for os_schedule_table in doc.getOs().getOsScheduleTableList():
            self.write_cell(sheet, row, 1, os_schedule_table.getName())
            self.write_cell(sheet, row, 2, os_schedule_table.getOsScheduleTableDuration(),
                            format={'alignment': Alignment(horizontal="center")})
            self.write_cell(sheet, row, 3, os_schedule_table.getOsScheduleTableRepeating(),
                            format={'alignment': Alignment(horizontal="center")})
            self.write_cell(sheet, row, 4, os_schedule_table.getOsScheduleTableCounterRef().getShortName(),
                            format={'alignment': Alignment(horizontal="center")})
            row += 1

            self.logger.debug("Write OsScheduleTable <%s>" % os_schedule_table.getName())

        self.auto_width(sheet)

    def write_os_counters(self, doc: EBModel):
        sheet = self.wb.create_sheet("OsCounter", 3)
        
        title_row = ["Name", "MaxAllowedValue", "MinCycle", "TicksPerBase", "Type", "SecondsPerTick"]
        self.write_title_row(sheet, title_row)

        row = 2
        for os_counter in doc.getOs().getOsCounterList():
            self.write_cell(sheet, row, 1, os_counter.getName())
            self.write_cell(sheet, row, 2, os_counter.getOsCounterMaxAllowedValue(),
                            format={'alignment': Alignment(horizontal="center")})
            self.write_cell(sheet, row, 3, os_counter.getOsCounterMinCycle(),
                            format={'alignment': Alignment(horizontal="center")})
            self.write_cell(sheet, row, 4, os_counter.getOsCounterTicksPerBase(),
                            format={'alignment': Alignment(horizontal="center")})
            self.write_cell(sheet, row, 5, os_counter.getOsCounterType(),
                            format={'alignment': Alignment(horizontal="center")})
            self.write_cell(sheet, row, 6, os_counter.getOsSecondsPerTick(),
                            format={'alignment': Alignment(horizontal="center")})
            row += 1

            self.logger.debug("Write OsScheduleTable <%s>" % os_counter.getName())

        self.auto_width(sheet)

    def write_expiry_points(self, doc: EBModel):
        sheet = self.wb.create_sheet("OsScheduleTableExpiryPoint", 4)

        title_row = ["ExpiryPoint", "OsScheduleTable", "OsCounter", "Offset (ms)", "Task"]
        self.write_title_row(sheet, title_row)

        row = 2
        for table in doc.getOs().getOsScheduleTableList():
            expiry_point_list = sorted(table.getOsScheduleTableExpiryPointList(),
                                       key=lambda o: o.getOsScheduleTblExpPointOffset())
            for expiry_point in expiry_point_list:
                self.write_cell(sheet, row, 1, expiry_point.getName())
                self.write_cell(sheet, row, 2, table.getName())
                self.write_cell(sheet, row, 3, table.getOsScheduleTableCounterRef().getShortName(),
                                format={'alignment': Alignment(horizontal="center")})
                self.write_cell(sheet, row, 4, expiry_point.getOsScheduleTblExpPointOffset(),
                                format={'alignment': Alignment(horizontal="center")})
                self.write_cell(sheet, row, 5, len(expiry_point.getOsScheduleTableTaskActivationList()),
                                format={'alignment': Alignment(horizontal="center")})
                row += 1

            self.logger.debug("Write OsScheduleTable <%s>" % table.getName())

        self.auto_width(sheet)

    def mk_memory_region_exists(self, doc: EBModel) -> bool:
        mk = doc.getOs().getOsMicrokernel()
        if mk is None:
            return False
        
        protection = mk.getMkMemoryProtection()
        if protection is None:
            return False
        
        if len(protection.getMkMemoryRegionList()) <= 0:
            return False
        
        return True
    
    def write_mk_memory_regions(self, doc: EBModel):
        if self.mk_memory_region_exists(doc) is True:
            sheet = self.wb.create_sheet("MkMemoryRegion", 5)

            title_row = [
                "Name", "Flags", "Initialize", "Global", "InitThread",
                "IdleThread", "OsThread", "ErrorHook", "ProtHook", "ShutdownHook",
                "Shutdown", "Kernel", "InitializePerCore"
            ]
            self.write_title_row(sheet, title_row)

            row = 2
            for region in doc.getOs().getOsMicrokernel().getMkMemoryProtection().getMkMemoryRegionList():
                self.write_cell(sheet, row, 1, region.getName())
                self.write_cell(sheet, row, 2, region.getMkMemoryRegionFlags(),
                                format={'alignment': Alignment(horizontal="center")})
                self.write_cell(sheet, row, 3, self.format_boolean(region.getMkMemoryRegionInitialize()),
                                format={'alignment': Alignment(horizontal="center")})
                self.write_cell(sheet, row, 4, self.format_boolean(region.getMkMemoryRegionGlobal()),
                                format={'alignment': Alignment(horizontal="center")})
                self.write_cell(sheet, row, 5, self.format_boolean(region.getMkMemoryRegionInitThreadAccess()),
                                format={'alignment': Alignment(horizontal="center")})
                self.write_cell(sheet, row, 6, self.format_boolean(region.getMkMemoryRegionIdleThreadAccess()),
                                format={'alignment': Alignment(horizontal="center")})
                self.write_cell(sheet, row, 7, self.format_boolean(region.getMkMemoryRegionOsThreadAccess()),
                                format={'alignment': Alignment(horizontal="center")})
                self.write_cell(sheet, row, 8, self.format_boolean(region.getMkMemoryRegionErrorHookAccess()),
                                format={'alignment': Alignment(horizontal="center")})
                self.write_cell(sheet, row, 9, self.format_boolean(region.getMkMemoryRegionProtHookAccess()),
                                format={'alignment': Alignment(horizontal="center")})
                self.write_cell(sheet, row, 10, self.format_boolean(region.getMkMemoryRegionShutdownHookAccess()),
                                format={'alignment': Alignment(horizontal="center")})
                self.write_cell(sheet, row, 11, self.format_boolean(region.getMkMemoryRegionShutdownAccess()),
                                format={'alignment': Alignment(horizontal="center")})
                self.write_cell(sheet, row, 12, self.format_boolean(region.getMkMemoryRegionKernelAccess()),
                                format={'alignment': Alignment(horizontal="center")})
                self.write_cell(sheet, row, 13, self.format_boolean(region.getMkMemoryRegionInitializePerCore()),
                                format={'alignment': Alignment(horizontal="center")})

                row += 1

                self.logger.debug("Write MkMemoryRegion <%s>" % region.getName())

            self.auto_width(sheet, {"B": 15})

    def write(self, filename, doc: EBModel, options={"skip_os_task": False}):
        self.logger.info("Writing <%s>" % filename)

        if not options['skip_os_task']:
            self.write_os_tasks(doc)
        self.write_os_isrs(doc)
        self.write_os_schedule_tables(doc)
        self.write_os_counters(doc)
        self.write_expiry_points(doc)
        self.write_mk_memory_regions(doc)

        self.save(filename)
