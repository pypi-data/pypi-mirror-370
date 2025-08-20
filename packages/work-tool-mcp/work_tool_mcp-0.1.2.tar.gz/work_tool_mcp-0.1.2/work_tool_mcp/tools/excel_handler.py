import mimetypes
import openpyxl
import time

from enum import Enum
from copy import copy
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

import re
import os

from .text_handler import TextHandler

class ExcelHandler:
    def __init__(self, filename: str):
        self.filename = filename

    def save_pdf_info_to_excel(self, data, is_font=False, is_color=False, is_size=False, is_space=False):
        # try:
            filename = self.filename
            # 创建Excel工作簿
            try:
                workbook = load_workbook(filename)
            except FileNotFoundError:
                workbook = Workbook()

            sheet = workbook.active
            if sheet is None:
                sheet = workbook.create_sheet(title="NewSheet")
            # 判断 sheet 是否为空表，为空表直接使用，不为空表则创建新表
            if not is_font:
                sheet = workbook.create_sheet()
            head_height = 30
            left_width = 16
            color_width = 10
            for file_info in [data]:
                current_row = 1
                filename = file_info["filename"]
                basename = file_info["basename"]
                
                sheet.title = basename
                # write into excel
                keys = ["文件名", "文件路径"]
                items = [basename, filename]
                
                if is_size:
                    sheet.title = "字号"
                    keys.append("使用字号")
                    size = sorted(file_info["size"], key=lambda unmber: float(unmber))
                    size_text = ", ".join(size)
                    items.append(size_text)
                elif is_font:
                    sheet.title = "字体"
                    keys.append("使用字体")
                    fonts = sorted(file_info["fonts"], key=lambda word: TextHandler.get_pinyin_first_letter(word))
                    fonts_text = ", ".join(fonts)
                    items.append(fonts_text)
                elif is_color:
                    sheet.title = "颜色"
                    keys.append("使用颜色")
                    colors = [f"({TextHandler.scale_and_round(str(tuple(item)))})" for item in file_info["colors"]]
                    colors = sorted(colors)
                    colors_text = ", ".join(colors)
                    items.append(colors_text)    
                elif is_space:
                    sheet.title = "空格"
                
                for index in range(len(items)):
                    row = current_row + index
                    self._set_sheet_value(keys[index], row, 1, sheet, width=left_width, height=head_height, bold=True, align_horizontal="center", background="76933C")
                    self._set_sheet_value(items[index], row, 2, sheet, width=left_width, height=head_height, background="D8E4BC")
                    start_cell = get_column_letter(2) + str(row)
                    end_cell = get_column_letter(4) + str(row)
                    sheet.merge_cells(f'{start_cell}:{end_cell}')
                current_row += len(items)
                
                for page_name, page_info in file_info["pages"].items():
                    pre_row = current_row
                    
                    if is_size:
                        # 使用字体
                        size_count = len(page_info["size_map"].items())
                        if size_count == 0:
                            self._set_sheet_value("", current_row, 2, sheet, width=left_width, bold=True, align_horizontal="center")
                            self._set_sheet_value("", current_row, 3, sheet)
                            start_cell = get_column_letter(3) + str(current_row)
                            end_cell = get_column_letter(4) + str(current_row)
                            sheet.merge_cells(f'{start_cell}:{end_cell}')
                            current_row += 1
                            
                        sorted_keys = sorted(page_info["size_map"].keys(), key=lambda key: float(key))
                        sorted_dict = {}
                        for key in sorted_keys:
                            sorted_dict[key] = page_info["size_map"][key]
                        for page_size, chars in sorted_dict.items():
                            chars_text = "".join(chars)
                            self._set_sheet_value(page_size, current_row, 2, sheet, width=left_width, bold=True)
                            self._set_sheet_value(chars_text, current_row, 3, sheet)
                            start_cell = get_column_letter(3) + str(current_row)
                            end_cell = get_column_letter(4) + str(current_row)
                            sheet.merge_cells(f'{start_cell}:{end_cell}')
                            current_row += 1
                    elif is_font:
                        # 使用字体
                        fonts_count = len(page_info["font_map"].items())
                        if fonts_count == 0:
                            self._set_sheet_value("", current_row, 2, sheet, width=left_width, bold=True, align_horizontal="center")
                            self._set_sheet_value("", current_row, 3, sheet)
                            start_cell = get_column_letter(3) + str(current_row)
                            end_cell = get_column_letter(4) + str(current_row)
                            sheet.merge_cells(f'{start_cell}:{end_cell}')
                            current_row += 1
                            
                        sorted_keys = sorted(page_info["font_map"].keys(), key=lambda key: TextHandler.get_pinyin_first_letter(key))
                        sorted_dict = {}
                        for key in sorted_keys:
                            sorted_dict[key] = page_info["font_map"][key]
                        for page_size, chars in sorted_dict.items():
                            chars_text = "".join(chars)
                            self._set_sheet_value(page_size, current_row, 2, sheet, width=left_width, bold=True)
                            self._set_sheet_value(chars_text, current_row, 3, sheet)
                            start_cell = get_column_letter(3) + str(current_row)
                            end_cell = get_column_letter(4) + str(current_row)
                            sheet.merge_cells(f'{start_cell}:{end_cell}')
                            current_row += 1
                    elif is_color:
                        # 使用颜色
                        colors_count = len(page_info["color_map"].items())
                        if colors_count == 0:
                            self._set_sheet_value("", current_row, 2, sheet, width=left_width)
                            self._set_sheet_value("", current_row, 3, sheet)
                            self._set_sheet_value("", current_row, 4, sheet, width=color_width)
                            current_row += 1
                        
                        sorted_keys = sorted(page_info["color_map"].keys(), key=lambda key: TextHandler.scale_and_round(key))
                        sorted_dict = {}
                        for key in sorted_keys:
                            sorted_dict[key] = page_info["color_map"][key]
                        for page_color, chars in sorted_dict.items():
                            background_color = TextHandler.cmyk_str_to_rgb(page_color)
                            text_color = TextHandler.scale_and_round(page_color)
                            colors_text = "".join(chars)
                            self._set_sheet_value(text_color, current_row, 2, sheet, width=left_width, bold=True)
                            self._set_sheet_value(colors_text, current_row, 3, sheet)
                            self._set_sheet_value("", current_row, 4, sheet, width=color_width, background=background_color)
                            current_row += 1
                    elif is_space:
                        for text in page_info["texts"]:
                            matches = [(m.start(0), m.end(0)) for m in re.finditer(r"\[空格\]", text)]
                            indics = [i for start, end in matches for i in range(start, end)]
                            self._set_sheet_rich_value(self._get_red_font(text, indics, color="595959"), current_row, 2, sheet)
                            start_cell = get_column_letter(2) + str(current_row)
                            end_cell = get_column_letter(4) + str(current_row)
                            sheet.merge_cells(f'{start_cell}:{end_cell}')
                            current_row += 1
                    
                    self._set_sheet_value(page_name, pre_row, 1, sheet, width=left_width, bold=True, align_horizontal="center", background="FABF8F")
                    if current_row - pre_row > 1:
                        start_cell = get_column_letter(1) + str(pre_row)
                        end_cell = get_column_letter(1) + str(current_row - 1)
                        sheet.merge_cells(f'{start_cell}:{end_cell}')
            # 保存工作簿
            workbook.save(self.filename)
            # 关闭 Excel 文件
            workbook.close()
            return True
        # except:
        #     return False
        
    
    def _set_sheet_rich_value(self, value: CellRichText, row_index: int, column_index: int, sheet, bold=False, width=60, height=60, background="ffffff", align_horizontal='left'):
        # 创建一个字体样式，字体为微软雅黑，大小为11
        font = Font(name='微软雅黑', size=11, bold=bold)
        alignment = Alignment(horizontal=align_horizontal, vertical='center', wrap_text=True)
        cell = sheet.cell(row=row_index, column=column_index, value=value)
        # 创建一个填充模式，设置背景色为蓝色
        fill = PatternFill(fill_type="solid", fgColor=background)
        # 创建边框样式
        border = Border(left=Side(border_style="thin"), 
                        right=Side(border_style="thin"), 
                        top=Side(border_style="thin"), 
                        bottom=Side(border_style="thin"))
        # 设置单元格的字体样式
        cell.font = font
        cell.alignment = alignment
        cell.fill = fill
        cell.border = border
        # 设置单元格的高度和宽度
        sheet.row_dimensions[row_index].height = width
        sheet.column_dimensions[sheet.cell(row=row_index, column=column_index).column_letter].width = height
    
    def _set_sheet_value(self, value, row_index: int, column_index: int, sheet, bold=False, width=60, height=60, color="000000", background="ffffff", align_horizontal='left'):
        # 创建一个字体样式，字体为微软雅黑，大小为11
        font = Font(name='微软雅黑', size=11, bold=bold, color=color)
        alignment = Alignment(horizontal=align_horizontal, vertical='center', wrap_text=True)
        cell = sheet.cell(row=row_index, column=column_index, value=value)
        # 创建一个填充模式，设置背景色为蓝色
        fill = PatternFill(fill_type="solid", fgColor=background)
        # 创建边框样式
        border = Border(left=Side(border_style="thin"), 
                        right=Side(border_style="thin"), 
                        top=Side(border_style="thin"), 
                        bottom=Side(border_style="thin"))
        # 设置单元格的字体样式
        cell.font = font
        cell.alignment = alignment
        cell.fill = fill
        cell.border = border
        # 设置单元格的高度和宽度
        sheet.row_dimensions[row_index].height = height
        sheet.column_dimensions[sheet.cell(row=row_index, column=column_index).column_letter].width = width

    def _get_red_font(self, text, red_marks, color="000000", tag_color="888888",tag = ""):
        red_font = InlineFont(color=Color("FF0000"), sz=12, b=True)
        black_font = InlineFont(color=Color(color), sz=12)
        grey_font = InlineFont(color=Color(tag_color), sz=12, b=True)
        
        sorted_indices = sorted(set(red_marks))  # 确保索引唯一且按升序排列
        rich_string = []
        last_pos = 0
        for index in sorted_indices:
            if index >= len(text):  # 确保索引有效
                continue
            # 添加红色之前的黑色文本
            if index > last_pos:
                rich_string.append(TextBlock(black_font, text[last_pos:index]))
            # 添加红色文本
            rich_string.append(TextBlock(red_font, text[index]))
            last_pos = index + 1

        # 添加最后一部分黑色文本
        if last_pos < len(text):
            rich_string.append(TextBlock(black_font, text[last_pos:]))
        if tag:
            rich_string.append(TextBlock(grey_font, f"\n{tag}"))
        return CellRichText(rich_string)
