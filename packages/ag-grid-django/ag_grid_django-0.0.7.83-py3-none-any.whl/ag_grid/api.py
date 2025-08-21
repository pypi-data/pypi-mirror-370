from datetime import datetime
import io
import json
import os

from django.apps import apps
from django.conf import settings
from django.db.models import (
    F,
    Q,
)
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.encoding import force_str
from django.utils.translation import gettext_lazy as _

from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework import status
from rest_framework.renderers import JSONRenderer
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.views import APIView as BaseAPIView

from .log import GridEditLog
from .permissions import AgGridModelPermission
from .registry import get_config, resource_registry

try:
    from ag_grid.contrib.notification.models import AgGridNotification
    from ag_grid.contrib.notification.utils import send_notification, send_model_notification

except ImportError:
    pass


# Field type mapping for converting Django model fields to AG Grid field types
FIELD_TYPE_MAP = {
    "AutoField": "number",
    "BigIntegerField": "number",
    "IntegerField": "number",
    "FloatField": "number",
    "DecimalField": "number",
    "CharField": "text",
    "TextField": "text",
    "EmailField": "text",
    "SlugField": "text",
    "BooleanField": "boolean",
    "DateField": "date",
    "DateTimeField": "datetime",
    "ForeignKey": "fk",
    "OneToOneField": "fk",
    "ManyToManyField": "m2m",
}

# Filter type mapping for AG Grid's column filters
FILTER_TYPE_MAP = {
    "AutoField": "agNumberColumnFilter",
    "BigIntegerField": "agNumberColumnFilter",
    "IntegerField": "agNumberColumnFilter",
    "FloatField": "agNumberColumnFilter",
    "DecimalField": "agNumberColumnFilter",
    "DateField": "agDateColumnFilter",
    "DateTimeField": "agDateColumnFilter",
    "ForeignKey": "",
    "OneToOneField": "",
    "ManyToManyField": "",
}

# Cell renderer mapping for displaying values in AG Grid
CELL_RENDERER_MAP = {
    "BooleanField": "agCheckboxCellRenderer",
    "DateField": "agDateCellRenderer",
    "DateTimeField": "agDateCellRenderer",
    "ForeignKey": "agTextCellRenderer",
    "OneToOneField": "agTextCellRenderer",
    "ManyToManyField": "agTextCellRenderer",
}

# Cell editor mapping for editing values in AG Grid
CELL_EDITOR_MAP = {
    "BooleanField": "agCheckboxCellRenderer",
    "DateField": "agDateCellEditor",
    "ForeignKey": "agSelectCellEditor",
}

# Cell editor parameter mapping for configuring editors
CELL_EDITOR_PARAM_MAP = {
    "ForeignKey": {
        "values": lambda field: [str(obj.id) for obj in field.related_model.objects.all()],
    },
}


class APIView(BaseAPIView):
    """
    Enhanced API view that provides improved permission handling
    """

    def dispatch(self, request, *args, **kwargs):
        """Override dispatch to ensure permissions are checked"""
        # Store kwargs for permission class to access
        self.kwargs = kwargs

        # Explicitly check permissions before proceeding
        for permission in self.get_permissions():
            if not permission.has_permission(request, self):
                return self.permission_denied(request, message=getattr(permission, "message", lambda r, v: None)(request, self))

        return super().dispatch(request, *args, **kwargs)

    def permission_denied(self, request, message=None, code=None):
        """
        Override permission_denied to provide clearer error messages
        """
        if message is None and hasattr(self.permission_classes[0], "message"):
            message = self.permission_classes[0]().message(request, self)

        response = {
            "error": "Permission denied",
            "detail": message or "You do not have permission to perform this action",
            "required_permissions": [f"{self.kwargs.get('app_label')}.{op}_{self.kwargs.get('model_name').lower()}" for op in ["view", "add", "change", "delete"]],
        }

        # Create a Response with a renderer explicitly set
        response_obj = Response(response, status=status.HTTP_403_FORBIDDEN)
        response_obj.accepted_renderer = JSONRenderer()
        response_obj.accepted_media_type = "application/json"
        response_obj.renderer_context = {
            "request": request,
            "view": self,
        }

        return response_obj


class AgGridHeaderAPIView(APIView):
    """
    API view for retrieving AG Grid column header configurations
    based on model metadata and configuration
    """

    permission_classes = [AgGridModelPermission]

    @swagger_auto_schema(
        operation_description=_("Get headers for AgGrid based on model configuration"),
        operation_summary=_("Get AgGrid Headers"),
        responses={
            200: openapi.Response(
                description="List of headers for the grid",
                schema=openapi.Schema(
                    type=openapi.TYPE_ARRAY,
                    items=openapi.Schema(
                        type=openapi.TYPE_OBJECT,
                        properties={
                            "field": openapi.Schema(type=openapi.TYPE_STRING, description="Field name"),
                            "headerName": openapi.Schema(type=openapi.TYPE_STRING, description="Header name"),
                            "selectionConfigs": openapi.Schema(type=openapi.TYPE_OBJECT, description="Selection configs dictionary"),
                            "editable": openapi.Schema(type=openapi.TYPE_BOOLEAN, description="Is field editable"),
                            "sortable": openapi.Schema(type=openapi.TYPE_BOOLEAN, description="Is field sortable"),
                            "pinned": openapi.Schema(type=openapi.TYPE_STRING, enum=["left", "right"], description="Is field pinned"),
                            "type": openapi.Schema(type=openapi.TYPE_STRING, description="Field type"),
                            "filter": openapi.Schema(type=openapi.TYPE_STRING, description="Filter type"),
                            "cellRenderer": openapi.Schema(type=openapi.TYPE_STRING, description="Cell renderer type"),
                            "cellEditor": openapi.Schema(type=openapi.TYPE_STRING, description="Cell editor type"),
                            "cellEditorParams": openapi.Schema(
                                type=openapi.TYPE_OBJECT,
                                description="Parameters for the cell editor",
                                additional_properties=openapi.Schema(type=openapi.TYPE_STRING),
                            ),
                        },
                    ),
                ),
            ),
            404: openapi.Response(description="Model or configuration not found", schema=openapi.Schema(type=openapi.TYPE_OBJECT, properties={"error": openapi.Schema(type=openapi.TYPE_STRING)})),
        },
        tags=[_("AgGrid")],
    )
    def get(self, request, app_label, model_name):
        """
        Get the column definitions for an AG Grid instance based on a Django model
        """
        try:
            try:
                model = apps.get_model(app_label, model_name)
            except LookupError:
                return Response({"error": "Model not found"}, status=status.HTTP_404_NOT_FOUND)

            config = get_config(model)

            if not config:
                headers = []

                model_fields = {f.name: f for f in model._meta.get_fields() if hasattr(f, "name")}

                field_list = [f.name for f in model._meta.fields]

                if "id" in field_list and field_list[0] != "id":
                    field_list.remove("id")
                    field_list.insert(0, "id")
                
                for field_name in field_list:
                    field = model_fields.get(field_name)
                    if field:
                        internal_type = field.get_internal_type()
                        field_type = FIELD_TYPE_MAP.get(internal_type, "text")
                        filter_type = FILTER_TYPE_MAP.get(internal_type, "agTextColumnFilter")
                        cell_renderer = CELL_RENDERER_MAP.get(internal_type, "agTextCellRenderer")
                        
                        editable = field_name != "id"
                        
                        headers.append({
                            "field": field_name,
                            "headerName": field.verbose_name.title() if hasattr(field, "verbose_name") else field_name.replace("_", " ").title(),
                            "selectionConfigs": {},
                            "editable": editable,
                            "sortable": True,
                            "pinned": "left" if field_name == "id" else None,
                            "type": field_type,
                            "filter": filter_type,
                            "cellRenderer": cell_renderer,
                            "enableCellTextSelection": True if field_type == "text" else False,
                            "cellStyle": { 'backgroundColor': '#f0f7ff' } if editable else {},
                        })
                
                return Response(headers)
            
            # Get the list of fields to display from the configuration
            field_list = config.get_list_display()

            # If the field list is already formatted as column definitions, return it directly
            if field_list and isinstance(field_list[0], dict):
                return Response(field_list)

            # Otherwise, build column definitions from the model fields
            if isinstance(field_list, (list, tuple)):
                field_list = list(field_list)
            else:
                field_list = [f.name for f in model._meta.fields]

            # Get custom header names if available
            custom_headers = {}
            if hasattr(config, "get_header_names") and callable(config.get_header_names):
                custom_headers = config.get_header_names()

            headers = []
            model_fields = {f.name: f for f in model._meta.get_fields() if hasattr(f, "name")}

            # Get selection configs if available
            selection_configs = {}
            if hasattr(config, "get_selection_configs") and callable(config.get_selection_configs):
                selection_configs = config.get_selection_configs()


            # Process each field in the field list
            for field_name in field_list:
                # Check if there's a custom header name for this field
                custom_header = custom_headers.get(field_name)

                #  Check if there's a custom labels for this field
                selection_config = selection_configs.get(field_name, {})

                # Handle regular fields
                if field_name in model_fields:
                    field = model_fields[field_name]
                    internal_type = field.get_internal_type()
                    field_type = FIELD_TYPE_MAP.get(internal_type, "text")
                    filter_type = FILTER_TYPE_MAP.get(internal_type, "agTextColumnFilter") if not selection_config.get("type") else "agSetColumnFilter"
                    cell_renderer = CELL_RENDERER_MAP.get(internal_type, "agTextCellRenderer")
                    cell_editor_type = CELL_EDITOR_MAP.get(internal_type, "agTextCellEditor")
                    cell_editor_params = CELL_EDITOR_PARAM_MAP.get(internal_type, {})

                    # Handle foreign key fields specially for dropdowns
                    if internal_type in ["ForeignKey", "OneToOneField"]:
                        if hasattr(config, "get_fk_display_field") and callable(config.get_fk_display_field):
                            display_field = config.get_fk_display_field(field.name)
                            if display_field:
                                # Get related model
                                related_model = field.related_model
                                # Create options using the display field
                                values = [None] if field.null else []
                                # Get all objects from related model and use the display field for values
                                objects = related_model.objects.all()
                                values.extend([str(getattr(obj, display_field)) for obj in objects])
                                # Set cell editor params with these values
                                cell_editor_params = {"values": values}

                                if selection_config.get("type") and not selection_config.get("labels"):
                                    id_display_pairs = [(str(obj.pk), str(getattr(obj, display_field))) for obj in objects]
                                    selection_config["labels"] = [display_val for _, display_val in id_display_pairs]
                            else:
                                # Fallback to using IDs if no display field is specified
                                related_model = field.related_model
                                values = [None] if field.null else []
                                values.extend([str(obj.id) for obj in related_model.objects.all()])
                                cell_editor_params = {"values": values}
                        else:
                            # Default behavior - get IDs from related model
                            related_model = field.related_model
                            values = [None] if field.null else []
                            values.extend([str(obj.id) for obj in related_model.objects.all()])
                            cell_editor_params = {"values": values}

                            if selection_config.get("type") and not selection_config.get("labels"):
                                objects = related_model.objects.all()
                                selection_config["labels"] = [str(obj.pk) for obj in objects]

                        cell_renderer = "agTextCellRenderer"

                    # Add the column definition
                    headers.append(
                        {
                            "field": field.name,
                            "headerName": custom_header or (field.verbose_name.title() if hasattr(field, "verbose_name") else field.name.replace("_", " ").title()),
                            "selectionConfigs": selection_config,
                            "editable": field.name in config.get_editable_fields(),
                            "sortable": field.name in config.get_sortable_fields(),
                            "pinned": "left" if field.name in config.get_left_pinning() else ("right" if field.name in config.get_right_pinning() else None),
                            "type": field_type,
                            "filter": filter_type,
                            "cellRenderer": cell_renderer,
                            "cellEditor": cell_editor_type,
                            "cellEditorParams": cell_editor_params,
                            "enableCellTextSelection": True if field_type == "text" else False,
                            "ensureDomOrder": True if field_type == "text" else False,
                            "cellStyle": { 'backgroundColor': '#f0f7ff' } if field.name in config.get_editable_fields() else {},
                        }
                    )

                # Handle related fields (those with "__")
                elif "__" in field_name:
                    parts = field_name.split("__")
                    relation_name = parts[0]
                    target_field = parts[1]

                    if relation_name in model_fields:
                        relation_field = model_fields[relation_name]

                        # Get the related model and field
                        if hasattr(relation_field, "related_model"):
                            related_model = relation_field.related_model
                            try:
                                related_field = related_model._meta.get_field(target_field)

                                # Create header for the related field
                                internal_type = related_field.get_internal_type()
                                field_type = FIELD_TYPE_MAP.get(internal_type, "text")
                                filter_type = FILTER_TYPE_MAP.get(internal_type, "agTextColumnFilter") if not selection_config.get("type") else "agSetColumnFilter"

                                # Selection config for related fields
                                if selection_config and selection_config.get("type"):
                                    if selection_config.get("labels"):
                                        selection_config["labels"] = [str(label) for label in selection_config["labels"]]
                                    else:
                                        selection_config["labels"] = [str(obj) for obj in related_model.objects.all()]

                                # Use custom header if available, otherwise use default
                                if custom_header:
                                    header_name = custom_header
                                elif hasattr(related_field, "verbose_name"):
                                    header_name = f"{relation_field.verbose_name} {related_field.verbose_name}".title()
                                else:
                                    header_name = field_name.replace("_", " ").title()

                                headers.append(
                                    {
                                        "field": field_name,
                                        "headerName": header_name,
                                        "selectionConfigs": selection_config,
                                        "editable": field_name in config.get_editable_fields(),
                                        "sortable": field_name in config.get_sortable_fields(),
                                        "type": field_type,
                                        "filter": filter_type,
                                        "cellRenderer": CELL_RENDERER_MAP.get(internal_type, "agTextCellRenderer"),
                                        "cellEditor": CELL_EDITOR_MAP.get(internal_type, "agTextCellEditor"),
                                        "cellEditorParams": CELL_EDITOR_PARAM_MAP.get(internal_type, {}),
                                        "enableCellTextSelection": True if field_type == "text" else False,
                                        "ensureDomOrder": True if field_type == "text" else False,
                                        "cellStyle": { 'backgroundColor': '#f0f7ff' } if field_name in config.get_editable_fields() else {},
                                    }
                                )
                            except:
                                # If related field can't be found, add a basic header
                                headers.append(
                                    {
                                        "field": field_name,
                                        "headerName": custom_header or field_name.replace("_", " ").title(),
                                        "selectionConfigs": selection_config,
                                        "editable": field_name in config.get_editable_fields(),
                                        "sortable": field_name in config.get_sortable_fields(),
                                        "type": "text",
                                        "filter": "agTextColumnFilter",
                                        "cellRenderer": "agTextCellRenderer",
                                        "cellEditor": "agTextCellEditor",
                                        "cellEditorParams": {},
                                        "enableCellTextSelection": True if field_type == "text" else False,
                                        "ensureDomOrder": True if field_type == "text" else False,
                                        "cellStyle": { 'backgroundColor': '#f0f7ff' } if field_name in config.get_editable_fields() else {},
                                    }
                                )
            print(f"Headers for {app_label}.{model_name}: {headers}")
            return Response(headers)
        except LookupError:
            return Response({"error": "Model not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            print(f"Error getting headers: {e}")
            return Response({"error": "Failed to retrieve headers"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class AgGridUpdateAPIView(APIView):
    """
    API view for updating a specific field of a model instance
    with real-time notification via WebSockets
    """

    permission_classes = [AgGridModelPermission]

    @swagger_auto_schema(
        operation_description=_("Update a specific field of a model instance"),
        operation_summary=_("Update Model Field"),
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=["field", "value"],
            properties={
                "field": openapi.Schema(type=openapi.TYPE_STRING, description="Field name to update"),
                "value": openapi.Schema(type=openapi.TYPE_STRING, description="New value for the field"),
                "previousValue": openapi.Schema(type=openapi.TYPE_STRING, description="Previous value of the field (optional)"),
            },
        ),
        responses={
            200: openapi.Response(description="Update successful", schema=openapi.Schema(type=openapi.TYPE_OBJECT, properties={"success": openapi.Schema(type=openapi.TYPE_BOOLEAN, default=True)})),
            400: openapi.Response(description="Invalid data provided", schema=openapi.Schema(type=openapi.TYPE_OBJECT, properties={"error": openapi.Schema(type=openapi.TYPE_STRING)})),
            404: openapi.Response(
                description="Object not found", schema=openapi.Schema(type=openapi.TYPE_OBJECT, properties={"error": openapi.Schema(type=openapi.TYPE_STRING, default="Object not found")})
            ),
            403: openapi.Response(description="Authentication required", schema=openapi.Schema(type=openapi.TYPE_OBJECT, properties={"detail": openapi.Schema(type=openapi.TYPE_STRING)})),
        },
        tags=[_("AgGrid")],
    )
    def patch(self, request, app_label, model_name, pk):
        """
        Update a specific field value and broadcast the change via WebSockets
        """
        try:
            model = apps.get_model(app_label, model_name)
            config = get_config(model)
            instance = model.objects.get(pk=pk)
        except Exception:
            return Response({"error": "Object not found"}, status=status.HTTP_404_NOT_FOUND)

        # Handle the new format with "field" and "value" keys
        field = request.data.get("field")
        value = request.data.get("value")
        previous_value = request.data.get("previousValue")

        print(f"Updating {app_label}.{model_name} (ID: {pk}) - Field: {field}, Old: {previous_value}, Value: {value}")

        if not field:
            return Response({"error": "Missing 'field' parameter"}, status=status.HTTP_400_BAD_REQUEST)

        editable_fields = config.get_editable_fields() if config else None
        
        if (config and (field in editable_fields or not editable_fields)) or not config:
            old_value = getattr(instance, field)

            # Optimistic locking check - make sure the field hasn't changed since the client loaded it
            if previous_value is not None and str(previous_value) != str(old_value):
                return Response({"error": "Previous value does not match current value", "oldValue": old_value}, status=status.HTTP_400_BAD_REQUEST)

            # Check if this field is a foreign key
            field_obj = model._meta.get_field(field)
            if field_obj.get_internal_type() in ["ForeignKey", "OneToOneField"]:
                # This is a foreign key field
                if value is not None and value != "":
                    try:
                        # Get the related model
                        related_model = field_obj.related_model
                        # Find the related object by ID
                        if hasattr(config, "get_fk_display_field") and callable(config.get_fk_display_field):
                            # Get the field to use for display (name, title, etc.)
                            custom_display = config.get_fk_display_field(field)
                            if custom_display:
                                # Look up by custom display field instead of pk
                                related_obj = related_model.objects.get(**{custom_display: value})
                            else:
                                # Fallback to lookup by pk
                                try:
                                    # Try to convert value to int for numeric PKs
                                    pk_value = int(value)
                                except (ValueError, TypeError):
                                    pk_value = value
                                related_obj = related_model.objects.get(pk=pk_value)
                        else:
                            # Default behavior - lookup by pk
                            try:
                                # Try to convert value to int for numeric PKs
                                pk_value = int(value)
                            except (ValueError, TypeError):
                                pk_value = value
                            related_obj = related_model.objects.get(pk=pk_value)
                        # Set the relationship
                        setattr(instance, field, related_obj)
                    except ValueError:
                        return Response({"error": f"Invalid format for foreign key ID: {value}"}, status=status.HTTP_400_BAD_REQUEST)
                    except related_model.DoesNotExist:
                        return Response({"error": f"Related object with ID {value} not found"}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    # Handle null case
                    setattr(instance, field, None)
            else:
                # Not a foreign key, set normally
                setattr(instance, field, value)

            if str(old_value) != str(value):
                # Log the update
                GridEditLog.log_update(
                    model_name=f"{app_label}.{model_name}", object_id=str(pk), field=field, old_value=str(old_value), new_value=str(value), user=request.user if request.user.is_authenticated else None
                )

                # Broadcast the update via WebSockets
                notification_data = {
                    "id": str(pk),
                    "field": field,
                    "previousValue": str(old_value),
                    "value": value,
                    "user_id": request.user.id if request.user.is_authenticated else None,
                    "username": request.user.username if request.user.is_authenticated else None,
                    "app_label": app_label,
                    "model_name": model_name,
                }

                # Send real-time update notification using the utility function
                try:
                    instance.save()
                    send_model_notification({
                        "type": "model_update",
                        "data": notification_data,
                        "group": f"{app_label}_{model_name}"
                    })
                except Exception as e:
                    print(f"Error sending notification: {e}")
                    return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
                
            instance.save()
            return Response({"success": True, "field": field, "old_value": str(old_value), "new_value": str(value)})
        else:
            return Response({"error": f"Field '{field}' is not editable"}, status=status.HTTP_400_BAD_REQUEST)


def _get_form_field_type(field):
    """
    Map Django field types to HTML form field types for dynamic form generation
    """
    internal_type = field.get_internal_type()
    if internal_type in ["CharField", "TextField", "SlugField", "EmailField", "URLField"]:
        return "textarea" if internal_type == "TextField" else "text"
    elif internal_type in ["IntegerField", "PositiveIntegerField", "PositiveSmallIntegerField", "SmallIntegerField", "BigIntegerField"]:
        return "number"
    elif internal_type in ["DecimalField", "FloatField"]:
        return "number"
    elif internal_type in ["BooleanField", "NullBooleanField"]:
        return "checkbox"
    elif internal_type in ["DateField"]:
        return "date"
    elif internal_type in ["DateTimeField"]:
        return "datetime-local"
    elif internal_type in ["TimeField"]:
        return "time"
    elif internal_type in ["ForeignKey", "OneToOneField"]:
        return "select"
    elif internal_type in ["ManyToManyField"]:
        return "multiselect"
    else:
        return "text"


class AgGridDeleteAPIView(APIView):
    """
    API view for deleting a model instance
    """

    permission_classes = [AgGridModelPermission]

    @swagger_auto_schema(
        operation_description=_("Delete a specific model instance"),
        operation_summary=_("Delete Model Instance"),
        responses={
            200: openapi.Response(
                description="Object deleted successfully",
                schema=openapi.Schema(type=openapi.TYPE_OBJECT, properties={"success": openapi.Schema(type=openapi.TYPE_BOOLEAN, default=True), "message": openapi.Schema(type=openapi.TYPE_STRING)}),
            ),
            404: openapi.Response(description="Object not found", schema=openapi.Schema(type=openapi.TYPE_OBJECT, properties={"error": openapi.Schema(type=openapi.TYPE_STRING)})),
            403: openapi.Response(description="Permission denied", schema=openapi.Schema(type=openapi.TYPE_OBJECT, properties={"error": openapi.Schema(type=openapi.TYPE_STRING)})),
        },
        tags=[_("AgGrid")],
    )
    def delete(self, request, app_label, model_name, pk):
        """
        Delete a model instance and log the deletion
        """
        try:
            model = apps.get_model(app_label, model_name)
            instance = model.objects.get(pk=pk)
        except LookupError:
            return Response({"error": "Model not found"}, status=status.HTTP_404_NOT_FOUND)
        except model.DoesNotExist:
            return Response({"error": "Object not found"}, status=status.HTTP_404_NOT_FOUND)

        # Save object data before deletion for logging
        try:
            # Get serializable data from the instance
            object_data = {}
            for field in model._meta.fields:
                field_name = field.name
                value = getattr(instance, field_name)
                object_data[field_name] = str(value)

            # Delete the instance
            instance.delete()

            # Log the deletion
            GridEditLog.log_delete(model_name=f"{app_label}.{model_name}", object_id=str(pk), user=request.user if request.user.is_authenticated else None, object_data=object_data)

            return Response({"success": True, "message": f"{model_name} deleted successfully"})

        except Exception as e:
            return Response({"error": f"Failed to delete {model_name}", "details": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class AgGridFormFieldsAPIView(APIView):
    """
    API view for getting form field metadata for a model for dynamic form generation
    """

    permission_classes = [AgGridModelPermission]

    @swagger_auto_schema(
        operation_description=_("Get form field requirements for a model"),
        operation_summary=_("Get Form Field Requirements"),
        responses={
            200: openapi.Response(
                description="Form field requirements",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "fields": openapi.Schema(
                            type=openapi.TYPE_ARRAY,
                            items=openapi.Schema(
                                type=openapi.TYPE_OBJECT,
                                properties={
                                    "name": openapi.Schema(type=openapi.TYPE_STRING),
                                    "type": openapi.Schema(type=openapi.TYPE_STRING),
                                    "label": openapi.Schema(type=openapi.TYPE_STRING),
                                    "required": openapi.Schema(type=openapi.TYPE_BOOLEAN),
                                    "placeholder": openapi.Schema(type=openapi.TYPE_STRING),
                                    "validation": openapi.Schema(type=openapi.TYPE_OBJECT),
                                    "options": openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Schema(type=openapi.TYPE_OBJECT)),
                                    "options_endpoint": openapi.Schema(type=openapi.TYPE_STRING),
                                },
                            ),
                        ),
                        "model_info": openapi.Schema(
                            type=openapi.TYPE_OBJECT,
                            properties={
                                "app_label": openapi.Schema(type=openapi.TYPE_STRING),
                                "model_name": openapi.Schema(type=openapi.TYPE_STRING),
                                "verbose_name": openapi.Schema(type=openapi.TYPE_STRING),
                            },
                        ),
                    },
                ),
            ),
            404: openapi.Response(description="Model or configuration not found", schema=openapi.Schema(type=openapi.TYPE_OBJECT, properties={"error": openapi.Schema(type=openapi.TYPE_STRING)})),
        },
        tags=[_("AgGrid")],
    )
    def get(self, request, app_label, model_name):
        """
        Get form field metadata for a model to support dynamic form generation
        """
        try:
            model = apps.get_model(app_label, model_name)
        except LookupError:
            return Response({"error": "Model not found"}, status=status.HTTP_404_NOT_FOUND)

        config = get_config(model)
        if not config:
            return Response({"error": "Grid config not found"}, status=status.HTTP_404_NOT_FOUND)

        # Get form fields from config or generate from model
        form_fields = {}
        if hasattr(config, "form_fields"):
            form_fields = config.form_fields
        else:
            # Generate basic form fields from model if not defined
            for field in model._meta.fields:
                if field.name == "id" or field.name.endswith("_ptr"):
                    continue

                field_config = {
                    "type": _get_form_field_type(field),
                    "label": field.verbose_name.title() if hasattr(field, "verbose_name") else field.name.replace("_", " ").title(),
                    "required": not field.blank,
                    "placeholder": f'Enter {field.verbose_name if hasattr(field, "verbose_name") else field.name}',
                }

                form_fields[field.name] = field_config

        # Convert to a list format with field names included
        fields_list = []
        for field_name, field_config in form_fields.items():
            field_data = field_config.copy()
            field_data["name"] = field_name
            fields_list.append(field_data)

        # Add model information
        model_info = {
            "app_label": app_label,
            "model_name": model_name,
            "verbose_name": model._meta.verbose_name.title(),
            "verbose_name_plural": model._meta.verbose_name_plural.title(),
            "create_url": f"/api/aggrid/{app_label}/{model_name}/create/",
            "update_url_template": f"/api/aggrid/{app_label}/{model_name}/{{id}}/update/",
            "delete_url_template": f"/api/aggrid/{app_label}/{model_name}/{{id}}/delete/",
        }

        return Response({"fields": fields_list, "model_info": model_info})

class AgGridFormForeignKeyOptionsAPIView(APIView):
    """
    API view for getting options for foreign key fields in a model
    """

    permission_classes = [AgGridModelPermission]

    @swagger_auto_schema(
        operation_description=_("Get options for foreign key fields in a model"),
        operation_summary=_("Get Foreign Key Options"),
        manual_parameters=[
            openapi.Parameter(
                "search",
                openapi.IN_QUERY,
                description=_("Search query to filter options"),
                type=openapi.TYPE_STRING,
            ),
        ],
        responses={
            200: openapi.Response(
                description="Foreign key options",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "options": openapi.Schema(
                            type=openapi.TYPE_ARRAY,
                            items=openapi.Schema(type=openapi.TYPE_OBJECT, properties={"id": openapi.Schema(type=openapi.TYPE_INTEGER), "name": openapi.Schema(type=openapi.TYPE_STRING)}),
                        )
                    },
                ),
            ),
            404: openapi.Response(description="Model or field not found", schema=openapi.Schema(type=openapi.TYPE_OBJECT, properties={"error": openapi.Schema(type=openapi.TYPE_STRING)})),
        },
        tags=[_("AgGrid")],
    )
    def get(self, request, app_label, model_name, field_name):
        """
        Get options for a foreign key field in a model
        """
        try:
            model = apps.get_model(app_label, model_name)
            field = model._meta.get_field(field_name)
        except LookupError:
            return Response({"error": "Model not found"}, status=status.HTTP_404_NOT_FOUND)
        except model.DoesNotExist:
            return Response({"error": "Field not found"}, status=status.HTTP_404_NOT_FOUND)

        if field.get_internal_type() not in ["ForeignKey", "OneToOneField"]:
            return Response({"error": f"Field '{field_name}' is not a foreign key"}, status=status.HTTP_400_BAD_REQUEST)

        related_model = field.related_model
        config = get_config(model)
        
        # Get all objects from the related model
        options = []
        
        # Check if config has form_fields with options configuration
        display_field = None
        if not config:
            return Response({"error": "Grid configuration not found"}, status=status.HTTP_404_NOT_FOUND)
        
        if not hasattr(config, 'form_fields'):
            return Response({"error": "form_fields configuration not found in grid config"}, status=status.HTTP_400_BAD_REQUEST)
        
        if field_name not in config.form_fields:
            return Response({"error": f"Field '{field_name}' not found in form_fields configuration"}, status=status.HTTP_400_BAD_REQUEST)
        
        field_config = config.form_fields[field_name]
        if not isinstance(field_config, dict) or 'options' not in field_config:
            return Response({"error": f"'options' configuration not found for field '{field_name}'"}, status=status.HTTP_400_BAD_REQUEST)
        
        options_config = field_config['options']
        if not isinstance(options_config, dict) or 'key' not in options_config:
            return Response({"error": f"'key' not specified in options configuration for field '{field_name}'"}, status=status.HTTP_400_BAD_REQUEST)
        
        display_field = options_config['key']
        
        # Validate that the display field exists in the related model
        try:
            related_model._meta.get_field(display_field)
        except:
            return Response({"error": f"Display field '{display_field}' not found in related model '{related_model.__name__}'"}, status=status.HTTP_400_BAD_REQUEST)
        
        search_query = request.query_params.get("search", None)
        filters = {}
        
        queryset = related_model.objects.all()

        if search_query:
            filters[f"{display_field}__icontains"] = search_query
            queryset = queryset.filter(**filters)
            queryset = queryset[:20]


        for obj in queryset:
            if hasattr(obj, display_field):
                # Use the specified display field
                display_value = getattr(obj, display_field)
            else:
                return Response({"error": f"Display field '{display_field}' not accessible on object"}, status=status.HTTP_400_BAD_REQUEST)
                
            option_data = {
                "id": obj.pk,
                "name": display_value,
            }
            options.append(option_data)

        return Response({"options": options})
        # """
        # Get options for a foreign key field in a model
        # """
        # try:
        #     model = apps.get_model(app_label, model_name)
        #     field = model._meta.get_field(field_name)
        # except LookupError:
        #     return Response({"error": "Model not found"}, status=status.HTTP_404_NOT_FOUND)
        # except model.DoesNotExist:
        #     return Response({"error": "Field not found"}, status=status.HTTP_404_NOT_FOUND)

        # if field.get_internal_type() not in ["ForeignKey", "OneToOneField"]:
        #     return Response({"error": f"Field '{field_name}' is not a foreign key"}, status=status.HTTP_400_BAD_REQUEST)

        # related_model = field.related_model
        # config = get_config(model)
        
        # # Get all objects from the related model
        # options = []
        
        # # Check if config has form_fields with options configuration
        # display_field = None
        # if config and hasattr(config, 'form_fields') and field_name in config.form_fields:
        #     field_config = config.form_fields[field_name]
        #     if isinstance(field_config, dict) and 'options' in field_config:
        #         options_config = field_config['options']
        #         if isinstance(options_config, dict) and 'key' in options_config:
        #             display_field = options_config['key']
        
        # # If no display field specified in form_fields, check for get_fk_display_field method
        # if not display_field and config and hasattr(config, "get_fk_display_field") and callable(config.get_fk_display_field):
        #     display_field = config.get_fk_display_field(field_name)
        
        # for obj in related_model.objects.all():
        #     if display_field and hasattr(obj, display_field):
        #         # Use the specified display field
        #         display_value = getattr(obj, display_field)
        #     else:
        #         # Fallback to string representation
        #         display_value = str(obj)
                
        #     option_data = {
        #         "id": obj.pk,
        #         "name": display_value,
        #     }
        #     options.append(option_data)

        # return Response({"options": options})

class AgGridFormCreateAPIView(APIView):
    
    permission_classes = [AgGridModelPermission]

    @swagger_auto_schema(
        operation_description=_("Create a new model instance using form fields configuration"),
        operation_summary=_("Create Model Instance"),
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            description="Form data based on form_fields configuration",
            additional_properties=True,
        ),
        responses={
            201: openapi.Response(
                description="Model instance created successfully",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "success": openapi.Schema(type=openapi.TYPE_BOOLEAN),
                        "message": openapi.Schema(type=openapi.TYPE_STRING),
                        "data": openapi.Schema(type=openapi.TYPE_OBJECT),
                        "id": openapi.Schema(type=openapi.TYPE_INTEGER),
                    },
                ),
            ),
            400: openapi.Response(
                description="Validation error",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        "success": openapi.Schema(type=openapi.TYPE_BOOLEAN),
                        "error": openapi.Schema(type=openapi.TYPE_STRING),
                        "errors": openapi.Schema(type=openapi.TYPE_OBJECT),
                    }
                )
            ),
            404: openapi.Response(description="Model not found", schema=openapi.Schema(type=openapi.TYPE_OBJECT, properties={"error": openapi.Schema(type=openapi.TYPE_STRING)})),
        },
        tags=[_("AgGrid")],
    )
    def post(self, request, app_label, model_name):
        """
        Create a new model instance using form_fields configuration
        """
        try:
            model = apps.get_model(app_label, model_name)
        except LookupError:
            return Response({"error": "Model not found"}, status=status.HTTP_404_NOT_FOUND)

        config = get_config(model)
        if not config:
            return Response({"error": "Grid configuration not found"}, status=status.HTTP_404_NOT_FOUND)
        
        if not hasattr(config, 'form_fields'):
            return Response({"error": "form_fields configuration not found in grid config"}, status=status.HTTP_400_BAD_REQUEST)

        # Validate and process form data
        form_data = json.loads(request.body)
        validated_data = {}
        errors = {}

        for field_name, field_config in config.form_fields.items():
            if not isinstance(field_config, dict):
                continue

            field_value = form_data[field_name]
            
            # Check required fields
            if field_config.get('required', False) and (field_value is None or field_value == ''):
                errors[field_name] = f"{field_config.get('label', field_name)} is required"
                continue

            # Skip validation if field is empty and not required
            if field_value is None or field_value == '':
                continue

            # Validate field based on type
            field_type = field_config.get('type', 'text')
            
            try:
                if field_type == 'select' or field_type == 'raw_id':
                    # Handle foreign key fields
                    django_field = model._meta.get_field(field_name)
                    if django_field.get_internal_type() in ["ForeignKey", "OneToOneField"]:
                        # Validate that the foreign key object exists
                        related_model = django_field.related_model
                        try:
                            related_obj = related_model.objects.get(pk=field_value)
                            validated_data[field_name] = related_obj
                        except related_model.DoesNotExist:
                            errors[field_name] = f"Invalid {field_config.get('label', field_name)} selection"
                    else:
                        validated_data[field_name] = int(field_value)

                elif field_type == 'text':
                    # Handle text fields with validation
                    validation = field_config.get('validation', {})
                    
                    # Check minimum length
                    min_length = validation.get('minLength', {})
                    if min_length and len(str(field_value)) < min_length.get('value', 0):
                        errors[field_name] = min_length.get('message', f"{field_config.get('label', field_name)} is too short")
                        continue
                    
                    # Check maximum length
                    max_length = validation.get('maxLength', {})
                    if max_length and len(str(field_value)) > max_length.get('value', 1000):
                        errors[field_name] = max_length.get('message', f"{field_config.get('label', field_name)} is too long")
                        continue
                    
                    validated_data[field_name] = str(field_value)
                
                elif field_type == 'number':
                    # Handle number fields
                    try:
                        validated_data[field_name] = float(field_value) if '.' in str(field_value) else int(field_value)
                    except (ValueError, TypeError):
                        errors[field_name] = f"{field_config.get('label', field_name)} must be a valid number"
                
                elif field_type == 'boolean':
                    # Handle boolean fields
                    validated_data[field_name] = bool(field_value)
                
                else:
                    # Default handling
                    validated_data[field_name] = field_value
                    
            except Exception as e:
                errors[field_name] = f"Error processing {field_config.get('label', field_name)}: {str(e)}"

        # Return validation errors if any
        if errors:
            return Response({
                "success": False,
                "error": "Validation failed",
                "errors": errors
            }, status=status.HTTP_400_BAD_REQUEST)

        # Create the model instance
        try:
            instance = model.objects.create(**validated_data)
            
            # Log the creation
            object_data = {field: str(getattr(instance, field)) for field in request.data.keys() if hasattr(instance, field)}
            GridEditLog.log_create(
                model_name=f"{app_label}.{model_name}",
                object_id=str(instance.pk),
                user=request.user if request.user.is_authenticated else None,
                object_data=object_data
            )

            # Prepare response data
            response_data = {}
            for field_name in config.form_fields.keys():
                if hasattr(instance, field_name):
                    field_value = getattr(instance, field_name)
                    # Handle foreign key fields
                    if hasattr(field_value, 'pk'):
                        response_data[field_name] = {
                            "id": field_value.pk,
                            "name": str(field_value)
                        }
                    else:
                        response_data[field_name] = field_value

            return Response({
                "success": True,
                "message": f"{model.__name__} created successfully",
                "data": response_data,
                "id": instance.pk
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({
                "success": False,
                "error": f"Failed to create {model.__name__}: {str(e)}"
            }, status=status.HTTP_400_BAD_REQUEST)
        

class AgGridFilteredListView(APIView):
    """
    Base class for creating filtered list views compatible with AG Grid.

    This view provides:
    - Complex filtering based on AG Grid filter models
    - Sorting support with multiple columns
    - Pagination for infinite scrolling
    - Support for annotations and calculated fields
    - Query optimization with select_related
    """
    permission_classes = [AgGridModelPermission]

    app_label = None
    model_name = None

    def get_model(self):
        """Get the model class from app_label and model_name"""
        if not (self.app_label and self.model_name):
            # Try to get from URL parameters
            self.app_label = self.kwargs.get("app_label")
            self.model_name = self.kwargs.get("model_name")

        if not (self.app_label and self.model_name):
            raise ValueError("Model information not provided. Set app_label and model_name or pass in URL.")

        return apps.get_model(self.app_label, self.model_name)

    def get_config(self):
        """Get the AG Grid configuration for the model"""
        model = self.get_model()
        return get_config(model)

    def get_base_queryset(self):
        """Get the base queryset to work with"""
        model = self.get_model()
        config = self.get_config()

        if not config:
            # If no config is found, return the default queryset
            return model.objects.all()

        # Use custom queryset method from config if available
        if hasattr(config, "get_queryset") and callable(config.get_queryset):
            return config.get_queryset(model)

        return model.objects.all()

    def get_field_types(self):
        """Get mapping of field names to their types, including related fields"""
        model = self.get_model()
        config = self.get_config()

        if not config:
            # If no config is found, return default field types
            field_types = {field.name: field.get_internal_type() for field in model._meta.fields}
            return field_types
        
        # Start with model fields
        field_types = {}
        for field in model._meta.fields:
            field_types[field.name] = field.get_internal_type()

            # Add related fields if the field is a relationship
            if field.get_internal_type() in ["ForeignKey", "OneToOneField"]:
                related_model = field.related_model
                for related_field in related_model._meta.fields:
                    field_types[f"{field.name}__{related_field.name}"] = related_field.get_internal_type()

        # Add custom field types from config if available
        if hasattr(config, "get_field_types") and callable(config.get_field_types):
            custom_types = config.get_field_types()
            field_types.update(custom_types)

        # Add any extra fields defined in config
        if hasattr(config, "get_extra_fields") and callable(config.get_extra_fields):
            for field_info in config.get_extra_fields():
                if "field" in field_info and "type" in field_info:
                    field_types[field_info["field"]] = field_info["type"]

        return field_types

    def apply_annotations(self, queryset):
        """Apply any annotations defined in the config"""
        config = self.get_config()

        if not config:
            # If no config is found, return the queryset as is
            return queryset

        if hasattr(config, "get_annotations") and callable(config.get_annotations):
            annotations = config.get_annotations(queryset)
            if annotations:
                queryset = queryset.annotate(**annotations)

        return queryset

    def apply_select_related(self, queryset):
        """Apply select_related based on fields used in list_display"""
        config = self.get_config()

        if not config:
            # If no config is found, return the queryset as is
            return queryset
        
        # Get list of fields to display
        field_list = config.get_list_display() if hasattr(config, "get_list_display") else []

        # Find relations to select_related
        relations = []
        for field in field_list:
            if isinstance(field, str) and "__" in field:
                relation = field.split("__")[0]
                if relation not in relations:
                    relations.append(relation)

        # Apply select_related
        if relations:
            queryset = queryset.select_related(*relations)

        return queryset

    def apply_filter(self, queryset, filter_params):
        """Apply filters from AG Grid filter model with relation support"""
        if not filter_params:
            return queryset

        filters = json.loads(filter_params)
        q_objects = Q()
        field_types = self.get_field_types()

        # Make sure all needed relations are selected
        relations_to_select = []
        for key in filters.keys():
            if "__" in key:
                relation = key.split("__")[0]
                if relation not in relations_to_select:
                    relations_to_select.append(relation)

        # Apply select_related for all relations used in filtering
        if relations_to_select:
            queryset = queryset.select_related(*relations_to_select)

        # Process each filter condition
        for key, filter_info in filters.items():
            # Skip empty filters
            if not filter_info:
                continue
            
            # Handle set filters (multi-select)
            if isinstance(filter_info, dict) and ("filterType" in filter_info and filter_info["filterType"] == "set"):
                q_objects &= self. _process_set_filter(key, filter_info)
            # Handle date filters
            elif isinstance(filter_info, dict) and ("filterType" in filter_info and filter_info["filterType"] == "date" or "dateFrom" in filter_info):
                q_objects &= self._process_date_filter(key, filter_info, field_types)
            # Handle number filters
            elif isinstance(filter_info, dict) and ("filterType" in filter_info and filter_info["filterType"] == "number"):
                q_objects &= self._process_number_filter(key, filter_info)
            # Handle text and other filters
            else:
                q_objects &= self._process_text_filter(key, filter_info)

        return queryset.filter(q_objects)
    
    def _process_set_filter(self, key, filter_info):
        # Extract values from the filter info
        values = filter_info.get("values", [])
        if not values:
            return Q()
        
        # Check if NULL values are included (None or "null" as string)
        null_included = None in values or "null" in values

        # Filter out real (non-null) values
        real_values = [v for v in values if v is not None and v != "null"]

        model = self.get_model()
        field_types = self.get_field_types()

        q_objects = Q()

        if key in field_types and field_types[key] in ["ForeignKey", "OneToOneField"]:
            field = model._meta.get_field(key)
            related_model = field.related_model
            config = self.get_config()

            display_field = None
            if config and hasattr(config, "get_fk_display_field") and callable(config.get_fk_display_field):
                display_field = config.get_fk_display_field(key)

            if display_field:
                pk_values = list(
                    related_model.objects
                    .filter(**{f"{display_field}__in": real_values})
                    .values_list("pk", flat=True)
                )

                if pk_values:
                    q_objects |= Q(**{f"{key}__in": pk_values})
                else:
                    return Q(pk__in=[])

        else:
            # Apply __in lookup for non-null values
            if real_values:
                q_objects |= Q(**{f"{key}__in": real_values})
            
        # Add __isnull condition if NULLs are included
        if null_included:
            q_objects |= Q(**{f"{key}__isnull": True})

        return q_objects

    def _process_date_filter(self, key, filter_info, field_types):
        """Process date filters from AG Grid"""
        # Composite filter with multiple conditions
        if "conditions" in filter_info:
            date_q = Q()
            operator = filter_info.get("operator", "AND")

            for condition in filter_info["conditions"]:
                condition_q = self._build_date_condition(key, condition, field_types)
                if condition_q:
                    if operator == "AND":
                        date_q &= condition_q
                    else:  # 'OR'
                        date_q |= condition_q

            return date_q
        # Single condition date filter
        else:
            return self._build_date_condition(key, filter_info, field_types)

    def _build_date_condition(self, key, condition, field_types):
        try:
            """Build a single date condition"""
            date_from = condition.get("dateFrom")
            date_to = condition.get("dateTo")
            filter_type = condition.get("type")

            # Handle null filters
            if filter_type == "blank":
                return Q(**{f"{key}__isnull": True})
            elif filter_type == "notBlank":
                return Q(**{f"{key}__isnull": False})

            if not (date_from or date_to):
                return None

            # Parse dates
            parsed_from = None
            parsed_to = None

            if date_from:
                try:
                    if " " in date_from:
                        # Handle datetime strings
                        parsed_from = datetime.strptime(date_from, "%Y-%m-%d %H:%M:%S")
                    else:
                        # Handle date strings
                        parsed_from = parse_date(date_from)
                except Exception as e:
                    print(f"Error parsing dateFrom: {e}")

            if date_to:
                try:
                    if " " in date_to:
                        # Handle datetime strings
                        parsed_to = datetime.strptime(date_to, "%Y-%m-%d %H:%M:%S")
                    else:
                        # Handle date strings
                        parsed_to = parse_date(date_to)
                except Exception as e:
                    print(f"Error parsing dateTo: {e}")

            # Determine field type (DateField or DateTimeField)
            field_type = field_types.get(key, "DateField")
            if field_type in ("datetime", "DateTimeField"):
                field_type = "DateTimeField"
            else:
                field_type = "DateField"

            # Build the query condition
            if filter_type == "equals":
                if field_type == "DateTimeField":
                    return Q(**{f"{key}__date": parsed_from})
                else:
                    return Q(**{f"{key}": parsed_from})
            elif filter_type == "notEqual":
                if field_type == "DateTimeField":
                    return ~Q(**{f"{key}__date": parsed_from})
                else:
                    return ~Q(**{f"{key}": parsed_from})
            elif filter_type == "lessThan":
                return Q(**{f"{key}__lt": parsed_from})
            elif filter_type == "greaterThan":
                return Q(**{f"{key}__gt": parsed_from})
            elif filter_type == "inRange" and parsed_to and parsed_from:
                # Make sure the dates are in the correct order
                if parsed_from > parsed_to:
                    parsed_from, parsed_to = parsed_to, parsed_from
                print(f"Date range filter: {key} between {parsed_from} and {parsed_to}")
                return Q(**{f"{key}__range": (parsed_from, parsed_to)})
            elif filter_type == "inRange" and parsed_from:
                # If only dateFrom is provided, use it as a lower bound
                return Q(**{f"{key}__gte": parsed_from})
            elif filter_type == "inRange" and parsed_to:
                # If only dateTo is provided, use it as an upper bound
                return Q(**{f"{key}__lte": parsed_to})

            return Q()  # Return empty Q if no match
        except Exception as e:
            print(f"Error processing date filter for {key}: {e}")
            return Q()  # Return empty Q object on error to avoid breaking the query

    def _process_number_filter(self, key, filter_info):
        """Process number filters from AG Grid"""
        # Handle composite filter with multiple conditions
        if "conditions" in filter_info:
            number_q = Q()
            operator = filter_info.get("operator", "AND")

            for condition in filter_info["conditions"]:
                # Process each condition individually
                condition_q = self._process_single_number_filter(key, condition)

                if operator == "AND":
                    number_q &= condition_q
                else:  # 'OR'
                    number_q |= condition_q

            return number_q
        # Single condition number filter
        else:
            return self._process_single_number_filter(key, filter_info)

    def _process_single_number_filter(self, key, filter_info):
        """Process a single number filter condition"""
        filter_type = filter_info.get("type")
        filter_value = filter_info.get("filter")

        # Handle null filters
        if filter_type == "blank":
            return Q(**{f"{key}__isnull": True})
        elif filter_type == "notBlank":
            return Q(**{f"{key}__isnull": False})

        if filter_value is None:
            return Q()

        # Try to convert to number if it's a string
        if isinstance(filter_value, str):
            try:
                filter_value = float(filter_value)
            except ValueError:
                pass

        # Apply the appropriate filter based on type
        if filter_type == "equals":
            return Q(**{f"{key}": filter_value})
        elif filter_type == "notEqual":
            return ~Q(**{f"{key}": filter_value})
        elif filter_type == "greaterThan":
            return Q(**{f"{key}__gt": filter_value})
        elif filter_type == "greaterThanOrEqual":
            return Q(**{f"{key}__gte": filter_value})
        elif filter_type == "lessThan":
            return Q(**{f"{key}__lt": filter_value})
        elif filter_type == "lessThanOrEqual":
            return Q(**{f"{key}__lte": filter_value})
        elif filter_type == "inRange":
            filter_to = filter_info.get("filterTo")
            if filter_to is not None:
                try:
                    if isinstance(filter_to, str):
                        filter_to = float(filter_to)
                    return Q(**{f"{key}__range": (filter_value, filter_to)})
                except ValueError:
                    pass

        return Q()

    def _process_text_filter(self, key, filter_info):
        """Process text filters from AG Grid"""
        filter_type = filter_info.get("type")
        filter_value = filter_info.get("filter")

        if filter_value is None:
            return Q()

        # Apply the appropriate filter based on type
        if filter_type == "equals":
            return Q(**{f"{key}": filter_value})
        elif filter_type == "notEqual":
            return ~Q(**{f"{key}": filter_value})
        elif filter_type == "contains":
            return Q(**{f"{key}__icontains": filter_value})
        elif filter_type == "notContains":
            return ~Q(**{f"{key}__icontains": filter_value})
        elif filter_type == "startsWith":
            return Q(**{f"{key}__istartswith": filter_value})
        elif filter_type == "endsWith":
            return Q(**{f"{key}__iendswith": filter_value})

        return Q()

    def apply_sort(self, queryset, sort_params):
        """Apply sorting from AG Grid sort model"""
        if not sort_params:
            return queryset

        sort_objects = json.loads(sort_params)
        sort_fields = []

        # Build list of sort fields
        for sort_object in sort_objects:
            col_id = sort_object.get("colId")
            sort_order = sort_object.get("sort")

            if not (col_id and sort_order):
                continue

            if sort_order == "asc":
                sort_fields.append(col_id)
            elif sort_order == "desc":
                sort_fields.append(f"-{col_id}")

        # Apply sorting
        if sort_fields:
            try:
                queryset = queryset.order_by(*sort_fields)
            except Exception as e:
                # Fallback to F expressions for complex field sorting
                for sort_object in sort_objects:
                    col_id = sort_object.get("colId")
                    sort_order = sort_object.get("sort")

                    if sort_order == "desc":
                        queryset = queryset.order_by(F(col_id).desc(nulls_last=True))
                    else:
                        queryset = queryset.order_by(F(col_id).asc(nulls_last=True))

        return queryset

    def get_field_list(self):
        """Get list of fields to include in response with FK display field support"""
        config = self.get_config()
        model = self.get_model()

        if not config:
            # If no config is found, return default fields
            field_list = [field.name for field in model._meta.fields]
            return field_list, {}

        # Get fields from list_display
        field_list = []
        if hasattr(config, "list_display"):
            field_list = list(config.list_display)
        elif hasattr(config, "get_list_display") and callable(config.get_list_display):
            field_list = config.get_list_display()

        # Always include 'id' field if it's not already there
        if "id" not in field_list:
            field_list.insert(0, "id")

        # Process foreign keys to use display fields instead of IDs
        model_fields = {f.name: f for f in model._meta.get_fields() if hasattr(f, "name")}
        fk_display_fields = {}
        additional_fields = []

        # First pass: identify foreign keys with display fields
        for field_name in list(field_list):
            if field_name in model_fields and model_fields[field_name].get_internal_type() in ["ForeignKey", "OneToOneField"]:
                # Check if a display field is specified
                if hasattr(config, "get_fk_display_field") and callable(config.get_fk_display_field):
                    display_field = config.get_fk_display_field(field_name)
                    if display_field:
                        # Store the display field name for this foreign key
                        fk_display_fields[field_name] = f"{field_name}__{display_field}"

                        # Add the display field to our field list
                        if fk_display_fields[field_name] not in field_list:
                            field_list.append(fk_display_fields[field_name])

            # Handle regular related fields
            elif "__" in field_name:
                base_field = field_name.split("__")[0]
                if base_field not in field_list and base_field not in additional_fields:
                    additional_fields.append(base_field)

        # Add necessary base fields
        field_list.extend(additional_fields)

        return field_list, fk_display_fields

    def dispatch(self, request, *args, **kwargs):
        """Ensure model info is available before processing the request"""
        if "app_label" not in kwargs and self.app_label:
            kwargs["app_label"] = self.app_label
        if "model_name" not in kwargs and self.model_name:
            kwargs["model_name"] = self.model_name

        if not (kwargs.get("app_label") and kwargs.get("model_name")):
            # Create Response with a renderer explicitly set
            response_obj = Response({"error": "Model information not provided"}, status=status.HTTP_400_BAD_REQUEST)
            response_obj.accepted_renderer = JSONRenderer()
            response_obj.accepted_media_type = "application/json"
            response_obj.renderer_context = {
                "request": request,
                "view": self,
            }
            return response_obj

        return super().dispatch(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_description="Get filtered list of records for AG Grid",
        operation_summary="Get AG Grid Filtered List",
        manual_parameters=[
            openapi.Parameter("startRow", openapi.IN_QUERY, description="Start row for pagination", type=openapi.TYPE_INTEGER),
            openapi.Parameter("endRow", openapi.IN_QUERY, description="End row for pagination", type=openapi.TYPE_INTEGER),
            openapi.Parameter("filter", openapi.IN_QUERY, description="AG Grid filter model (JSON)", type=openapi.TYPE_STRING),
            openapi.Parameter("sort", openapi.IN_QUERY, description="AG Grid sort model (JSON)", type=openapi.TYPE_STRING),
        ],
        responses={
            200: openapi.Response(
                description="Successful operation",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={"rows": openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Schema(type=openapi.TYPE_OBJECT)), "totalRows": openapi.Schema(type=openapi.TYPE_INTEGER)},
                ),
            ),
            400: "Bad request",
            403: "Permission denied",
            404: "Model not found",
        },
        tags=[_("AgGrid")],
    )
    def get(self, request, *args, **kwargs):
        """
        Process GET request with filtering, sorting and pagination
        for AG Grid infinite scrolling
        """
        try:
            # Get pagination parameters
            start_row = int(request.GET.get("startRow", 0))
            end_row = int(request.GET.get("endRow", 100))

            # Get filter and sort parameters
            filter_params = request.GET.get("filter")
            sort_params = request.GET.get("sort")

            # Get base queryset
            queryset = self.get_base_queryset()

            # Apply select_related for optimization
            queryset = self.apply_select_related(queryset)

            # Apply annotations
            queryset = self.apply_annotations(queryset)

            # Apply filters
            queryset = self.apply_filter(queryset, filter_params)

            # Apply custom filters
            queryset = self.apply_custom_filters(queryset, request)

            # Get total count before pagination
            total_rows = queryset.count()

            # Apply sorting
            queryset = self.apply_sort(queryset, sort_params)

            # Get field list
            field_list, fk_display_fields = self.get_field_list()

            # Apply pagination and convert to list
            rows = list(queryset.values(*field_list)[start_row:end_row])

            # Process foreign key display fields
            for row in rows:
                for fk_field, display_field in fk_display_fields.items():
                    if display_field in row and row[display_field] is not None:
                        # Replace the FK ID with its display value
                        row[fk_field] = row[display_field]

            return Response({"rows": rows, "totalRows": total_rows})

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def apply_custom_filters(self, queryset, request):
        """
        Override this method to add custom filters.
        This is called after standard AG Grid filters are applied.
        """
        return queryset


class AgGridExcelExportAPIView(APIView):
    """
    API view for exporting AG Grid data to Excel.

    This view supports:
    - Export selected rows by IDs
    - Configurable columns through Resource classes
    - Custom formatting
    """

    permission_classes = [AgGridModelPermission]

    @swagger_auto_schema(
        operation_description=_("Export AG Grid data to Excel format"),
        operation_summary=_("Export AG Grid Data to Excel"),
        manual_parameters=[
            openapi.Parameter("app_label", openapi.IN_PATH, description="Application label of the model", type=openapi.TYPE_STRING),
            openapi.Parameter("model_name", openapi.IN_PATH, description="Model name to export data from", type=openapi.TYPE_STRING),
        ],
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                "export_all_filtered": openapi.Schema(type=openapi.TYPE_BOOLEAN, description="Export all filtered items instead of selected rows", default=False),
                "selected_rows": openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Schema(type=openapi.TYPE_INTEGER), description="List of selected row IDs to export"),
                "filename": openapi.Schema(type=openapi.TYPE_STRING, description="Custom filename for the exported Excel file", default="export.xlsx"),
            },
        ),
        responses={
            200: openapi.Response(
                description="Excel file exported successfully",
                schema=openapi.Schema(type=openapi.TYPE_FILE, description="Excel file containing the exported data"),
            ),
            400: openapi.Response(description="Bad request", schema=openapi.Schema(type=openapi.TYPE_OBJECT, properties={"error": openapi.Schema(type=openapi.TYPE_STRING)})),
            403: openapi.Response(description="Permission denied", schema=openapi.Schema(type=openapi.TYPE_OBJECT, properties={"error": openapi.Schema(type=openapi.TYPE_STRING)})),
            404: openapi.Response(description="Model not found", schema=openapi.Schema(type=openapi.TYPE_OBJECT, properties={"error": openapi.Schema(type=openapi.TYPE_STRING)})),
        },
        tags=[_("AgGrid")],
    )
    def post(self, request, app_label, model_name):
        """
        Export AG Grid data to Excel format based on selected row IDs and registered resource
        """
        try:
            # Get model information from URL parameters
            model = apps.get_model(app_label, model_name)

            # Check if exporting all filtered items or just selected rows
            export_all_filtered = request.data.get("export_all_filtered", False)
            safe_ordering = []  # Holds valid field names for sorting (fallback)
            ordering = []  # Holds original requested sort fields

            selected_rows = request.data.get("selected_rows", [])

            if export_all_filtered:
                # Get filter and sort models from request
                filter_model = request.data.get("filter_model", {})
                sort_model = request.data.get("sort_model", [])

                # Start with all records and then apply filters
                queryset = model.objects.all()

                # Apply filters from AG Grid filter model
                if filter_model:
                    # 1. Convert filter model to JSON string
                    filter_json = json.dumps(filter_model)

                    # 2. Create FilteredListView instance and set required attributes
                    filter_view = AgGridFilteredListView()
                    filter_view.app_label = app_label
                    filter_view.model_name = model_name
                    filter_view.kwargs = {"app_label": app_label, "model_name": model_name}

                    # 3. Apply filtering using instance method
                    queryset = filter_view.apply_filter(queryset, filter_json)

                    # Apply sorting from AG Grid sort model
                if sort_model:
                    for sort in sort_model:
                        # Skip items without column ID
                        if not sort.get("colId"):
                            continue

                        # Handle ascending/descending sort
                        if sort.get("sort") == "desc":
                            ordering.append(f"-{sort.get('colId')}")  # Prefix with minus for descending
                        else:
                            ordering.append(sort.get("colId"))

                    # Apply sorting if we have valid sort fields
                    if ordering:
                        try:
                            # Try to apply all requested sort fields
                            queryset = queryset.order_by(*ordering)
                        except Exception as e:
                            # Handle sorting errors
                            print(f"Sorting error: {str(e)}")

                            # Get list of valid model fields
                            valid_fields = [f.name for f in model._meta.fields]
                            print(f"Valid model fields: {valid_fields}")

                            # Build a list of safe field names that actually exist
                            for field in ordering:
                                field_name = field.replace("-", "")  # Remove minus sign for field lookup
                                if field_name in valid_fields:
                                    safe_ordering.append(field)

                            # Try sorting again with valid fields only
                            if safe_ordering:
                                print(f"Trying with safe ordering: {safe_ordering}")
                                queryset = queryset.order_by(*safe_ordering)

                # Get only IDs to avoid memory issues with large datasets
                selected_rows = list(queryset.values_list("id", flat=True))

            # Set filename from request or use default
            filename = request.data.get("filename", f"{model_name}_export.xlsx")

            # Ensure filename has .xlsx extension
            if not filename.endswith(".xlsx"):
                filename += ".xlsx"

            # Require at least one row to export
            if not selected_rows:
                return Response({"error": "No rows selected for export"}, status=status.HTTP_400_BAD_REQUEST)

            # Get registered resource for export configuration
            resource = resource_registry.get_resource(model)
            if resource:
                # Use resource class for export formatting
                queryset = resource.get_queryset().filter(pk__in=selected_rows)

                # Apply ordering if specified
                if ordering:
                    queryset = queryset.order_by(*ordering)

                # Get export data using resource methods
                headers = resource.get_export_headers()
                data = resource.get_export_data(queryset)
                field_formats = resource.get_field_formats()
                sheet_title = force_str(model._meta.verbose_name_plural)[:31]
            else:
                # Get queryset based on selected rows
                queryset = model.objects.filter(pk__in=selected_rows)
                
                print(ordering)
                if ordering:
                    queryset = queryset.order_by(*ordering)

                # Get basic fields from model
                basic_fields = [f.name for f in model._meta.fields]
                
                # Ensure 'id' is always included in basic fields
                if 'id' not in basic_fields:
                    basic_fields.insert(0, 'id')
                
                # Generate headers from basic fields
                headers = []
                for field_name in basic_fields:
                    field = model._meta.get_field(field_name)
                    if hasattr(field, 'verbose_name'):
                        headers.append(field.verbose_name.title())
                    else:
                        headers.append(field_name.replace('_', ' ').title())
                
                # Get export data using default method
                data = self._get_default_export_data(queryset, basic_fields)
                
                field_formats = {}
                sheet_title = force_str(model._meta.verbose_name_plural)[:31]

            # Create Excel workbook
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = sheet_title

            # Set header row with bold formatting
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            # Write headers to first row
            for col_idx, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Write data rows
            if resource:
                # Resource-based data format (list of dicts)
                for row_idx, row_data in enumerate(data, 2):
                    for col_idx, field_name in enumerate(resource._meta["export_order"], 1):
                        if field_name in row_data:
                            value = row_data[field_name]

                            # Format special value types
                            if hasattr(value, "isoformat"):  # For date/datetime objects
                                value = value.isoformat()
                            elif hasattr(value, "__str__") and not isinstance(value, (str, int, float, bool, type(None))):
                                value = str(value)

                            sheet.cell(row=row_idx, column=col_idx, value=value)
            else:
                # Default data format (dictionary with field:value pairs)
                for row_idx, row_data in enumerate(data, 2):
                    for col_idx, field_name in enumerate(basic_fields, 1):
                        value = row_data.get(field_name, None)
                        
                        # Format special value types
                        if hasattr(value, "isoformat"):  # For date/datetime objects
                            value = value.isoformat()
                        elif hasattr(value, "__str__") and not isinstance(value, (str, int, float, bool, type(None))):
                            value = str(value)
                            
                        sheet.cell(row=row_idx, column=col_idx, value=value)

            # Apply column formats and widths
            for col_idx, field_name in enumerate(resource._meta["export_order"] if resource else basic_fields, 1):
                col_letter = get_column_letter(col_idx)

                # Set column width if specified in field formats
                if field_name in field_formats and isinstance(field_formats[field_name], dict) and "width" in field_formats[field_name]:
                    sheet.column_dimensions[col_letter].width = field_formats[field_name]["width"]
                else:
                    # Auto-adjust column width based on content
                    max_length = 0
                    for row_idx in range(1, len(data) + 2):  # +2 for header and 1-based indexing
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = max_length + 2  # Adding padding
                    sheet.column_dimensions[col_letter].width = min(adjusted_width, 50)  # Cap width at 50

            export_dir = os.path.join(settings.MEDIA_ROOT, "excel_exports", app_label, model_name)
            os.makedirs(export_dir, exist_ok=True)

            # Add timestamp to filename to prevent overwriting
            timestamp = timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M%S")
            unique_filename = f"{os.path.splitext(filename)[0]}_{timestamp}.xlsx"
            file_path = os.path.join(export_dir, unique_filename)

            # Save the workbook to the file
            buffer = io.BytesIO()
            workbook.save(buffer)
            buffer.seek(0)

            if "ag_grid.contrib.notification" in settings.INSTALLED_APPS:
                with open(file_path, "wb") as f:
                    f.write(buffer.getvalue())

                # Create download URL for the file
                relative_path = os.path.join("excel_exports", app_label, model_name, unique_filename)
                download_url = f"/media/{relative_path.replace(os.sep, '/')}"

                # Log the Excel export action
                GridEditLog.log_action(
                    model_name=f"{app_label}.{model_name}_excel_export",
                    object_id=None,
                    user=request.user if request.user.is_authenticated else None,
                    object_data={
                        "filename": force_str(unique_filename),
                        "record_count": len(selected_rows),
                        "export_type": force_str("filtered" if export_all_filtered else "selected"),
                        "filter_model": force_str(filter_model) if export_all_filtered and filter_model else None,
                        "sort_model": force_str(sort_model) if export_all_filtered and sort_model else None,
                        "file_path": force_str(file_path),
                        "download_url": force_str(download_url),
                    },
                )

                if request.user.is_authenticated:
                    # Create notification
                    notification = AgGridNotification.create_export_notification(user=request.user, filename=force_str(filename), file_url=force_str(download_url))

                    # Send real-time notification
                    send_notification(notification)

                # Return success response
                return Response(
                    {
                        "success": True,
                        "message": "Excel export is being processed. You will be notified when it's ready.",
                        "filename": force_str(unique_filename),
                    },
                    status=status.HTTP_202_ACCEPTED,
                )

            # Create HTTP response with Excel file
            response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = f'attachment; filename="{unique_filename}"'

            # Save the workbook to the response
            response.write(buffer.getvalue())

            # Log the Excel export action
            GridEditLog.log_action(
                model_name=f"{app_label}.{model_name}_excel_export",
                object_id=None,  # No specific object ID for exports
                user=request.user if request.user.is_authenticated else None,
                object_data={
                    "filename": force_str(unique_filename),
                    "record_count": len(selected_rows),
                    "export_type": "filtered" if export_all_filtered else "selected",
                    "filter_model": str(filter_model) if export_all_filtered and filter_model else None,
                    "sort_model": str(sort_model) if export_all_filtered and sort_model else None,
                },
            )

            return response

        except LookupError:
            return Response({"error": "Model not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            print(f"Export error: {str(e)}")
            return Response({"error": f"Failed to export data: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

    def _get_default_export_data(self, queryset, field_list):
        """Generate default export data when no resource is available"""
        data = []
        for obj in queryset:
            row_data = {}
            for field in field_list:
                try:
                    if "__" in field:
                        # Handle related fields by traversing relationships
                        parts = field.split("__")
                        value = obj
                        for part in parts:
                            if value is None:
                                break
                            value = getattr(value, part, None)
                    else:
                        # Handle regular fields
                        value = getattr(obj, field, None)

                    # Format special value types
                    if hasattr(value, "isoformat"):  # For date/datetime objects
                        value = value.isoformat()
                    elif hasattr(value, "__str__") and not isinstance(value, (str, int, float, bool, type(None))):
                        value = str(value)

                    row_data[field] = value
                except Exception as e:
                    print(f"Error getting field {field} for object {obj.pk}: {e}")
                    row_data[field] = None
            data.append(row_data)
        return data

    def _apply_select_related(self, queryset, config):
        """Optimize query by applying select_related for related fields"""
        field_list = config.get_list_display() if hasattr(config, "get_list_display") and callable(config.get_list_display) else []

        # Find relations to select_related
        relations = []
        for field in field_list:
            if isinstance(field, str) and "__" in field:
                relation = field.split("__")[0]
                if relation not in relations:
                    relations.append(relation)

        # Apply select_related to reduce number of queries
        if relations:
            queryset = queryset.select_related(*relations)

        return queryset


    def _get_field_headers(self, config, model):
        """모델의 기본 필드와 헤더 이름 가져오기"""
        # check if config has use_basic_fields_for_export attribute
        use_config_fields = False
        
        if hasattr(config, 'use_basic_fields_for_export') and config.use_basic_fields_for_export is False:
            use_config_fields = True
        
        if use_config_fields:
            # config has list_display or get_list_display
            field_list = []
            if hasattr(config, "list_display"):
                field_list = list(config.list_display)
            elif hasattr(config, "get_list_display") and callable(config.get_list_display):
                field_list = config.get_list_display()
        else:
            # get fields from model
            field_list = [f.name for f in model._meta.fields]
        
        # generate headers based on field names
        headers = []
        for field_name in field_list:
            field = next((f for f in model._meta.fields if f.name == field_name), None)
            if field and hasattr(field, 'verbose_name'):
                headers.append(field.verbose_name.title())
            else:
                headers.append(field_name.replace('_', ' ').title())
        
        return field_list, headers
    
