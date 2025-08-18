import xlwt
import openpyxl
import os

class excel_log:
    def __init__(self,filepath):
        self.Workbook = xlwt.Workbook()
        self.sheet=[]
        self.file = filepath
        
    def add_sheet(self,name):
        self.sheet.append(self.Workbook.add_sheet(name))
        
    def add_data(self,sheetnun,x,y,c):
        self.sheet[sheetnun].write(x,y,c)
    
    def save_sheet(self):
        self.Workbook.save(self.file)
        
class excel_log1:
    def __init__(self,filepath):
        self.filepath = filepath
        self.wb = openpyxl.Workbook()
        if os.path.exists(filepath) ==True:
            self.wb = openpyxl.load_workbook(filepath)
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb.get_sheet_by_name("Sheet"))
            
    def add_sheet(self,name):
        self.wb.create_sheet(name)
        
    def add_data(self,sheetname,x,y,c):
        ws = self.wb.get_sheet_by_name(sheetname)
        ws.cell(row=x+1,column=y+1,value=c)
        
    def save_sheet(self):
        self.wb.save(self.filepath)
        
    def change_data(self,sheetname,x,y,c):
        self.wb = openpyxl.load_workbook(self.filepath)
        self.add_data(sheetname,x,y,c)
        self.save_sheet()
    
    def get_data(self,sheetname,x,y):
        self.wb = openpyxl.load_workbook(self.filepath)
        ws = self.wb.get_sheet_by_name(sheetname)
        return ws.cell(row=x+1,column=y+1).value
        
# log = excel_log1("test.xlsx")
# log.add_sheet("#1")
# for i in range(0,10):
#     log.add_data("#1",i,0,i)
# log.save_sheet()