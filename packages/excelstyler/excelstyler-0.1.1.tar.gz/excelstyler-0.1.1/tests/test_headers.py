from io import BytesIO

import openpyxl
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from excelstyler.headers import create_header
from oauth2_provider.contrib.rest_framework import TokenHasReadWriteScope
from openpyxl import Workbook
from rest_framework.decorators import api_view, permission_classes

from excelstyler.values import create_value
from excelstyler.headers import create_header_freez


def test_header_creation():
    wb = openpyxl.Workbook()
    ws = wb.active
    create_header(ws, ["A", "B"], 1, 1)
    assert ws.cell(1, 1).value == "A"


@api_view(["POST"])
@permission_classes([TokenHasReadWriteScope])
@csrf_exempt
def test_cold_house_excel(request):
    """
    A simplified example Excel report for Cold Houses.
    Excel output support Persian name.
    """

    # --- Excel Setup ---
    output = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    workbook.remove(worksheet)
    worksheet = workbook.create_sheet("Cold House Info")
    worksheet.sheet_view.rightToLeft = True
    worksheet.insert_rows(1)

    # --- Header ---
    header = [
        'Row', 'Cold House Name', 'City', 'Address',
        'Total Weight', 'Allocated Weight', 'Remaining Weight',
        'Status', 'Broadcast', 'Relocate', 'Capacity'
    ]
    create_header_freez(worksheet, header, start_col=1, row=2, header_row=3)

    # --- Example Data ---
    # Here we use some mock data for testing
    example_data = [
        {
            'name': 'Cold House A',
            'city': 'Tehran',
            'address': 'Street 1',
            'total_input_weight': 1000,
            'total_allocated_weight': 700,
            'total_remain_weight': 300,
            'status': True,
            'broadcast': False,
            'relocate': True,
            'capacity': 1200
        },
        {
            'name': 'Cold House B',
            'city': 'Shiraz',
            'address': 'Street 2',
            'total_input_weight': 800,
            'total_allocated_weight': 500,
            'total_remain_weight': 300,
            'status': False,
            'broadcast': True,
            'relocate': False,
            'capacity': 1000
        }
    ]

    # --- Fill Data ---
    row_index = 3
    for i, house in enumerate(example_data, start=1):
        row_index += 1
        values = [
            i,
            house['name'],
            house['city'],
            house['address'],
            house['total_input_weight'],
            house['total_allocated_weight'],
            house['total_remain_weight'],
            'Active' if house['status'] else 'Inactive',
            'Yes' if house['broadcast'] else 'No',
            'Yes' if house['relocate'] else 'No',
            house['capacity']
        ]
        create_value(worksheet, values, start_col=row_index, row=1)

    # --- Save and Response ---
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="ColdHouseExample.xlsx"'
    response.write(output.getvalue())
    return response
