from openpyxl.styles import PatternFill, Alignment, Font

# Predefined cell styles for Excel
blue_fill = PatternFill(start_color="277358", fill_type="solid")  # Default blue fill
Alignment_CELL = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Center alignment with wrap
red_font = Font(color="C00000", bold=True)  # Bold red font

# Predefined color fills
GREEN_CELL = PatternFill(start_color="00B050", fill_type="solid")
RED_CELL = PatternFill(start_color="FCDFDC", fill_type="solid")
YELLOW_CELL = PatternFill(start_color="FFFF00", fill_type="solid")
ORANGE_CELL = PatternFill(start_color="FFC000", fill_type="solid")
BLUE_CELL = PatternFill(start_color="538DD5", fill_type="solid")
LIGHT_GREEN_CELL = PatternFill(start_color="92D050", fill_type="solid")
VERY_LIGHT_GREEN_CELL = PatternFill(start_color="5AFC56", fill_type="solid")
GRAY_CELL = PatternFill(start_color="B0B0B0", fill_type="solid")
CREAM_CELL = PatternFill(start_color="D8AA72", fill_type="solid")
LIGHT_CREAM_CELL = PatternFill(start_color="E8C6A0", fill_type="solid")
VERY_LIGHT_CREAM_CELL = PatternFill(start_color="FAF0E7", fill_type="solid")

# Dictionary to map color names to PatternFill objects
color_dict = {
    'green': GREEN_CELL,
    'yellow': YELLOW_CELL,
    'blue': BLUE_CELL,
    'red': RED_CELL,
    'light_green': LIGHT_GREEN_CELL,
    'very_light_green': VERY_LIGHT_GREEN_CELL,
    'gray': GRAY_CELL
}
