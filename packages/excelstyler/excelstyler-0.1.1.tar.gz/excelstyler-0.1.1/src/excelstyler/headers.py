import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from .styles import *


def create_header(
        worksheet, data, start_col, row, height=None, width=None, color=None, text_color=None, border_style=None
):
    """
    Create a header row in an Excel worksheet with optional styling.

    Parameters:
    -----------
    worksheet : openpyxl.worksheet.worksheet.Worksheet
        The worksheet where the header will be created.
    data : list
        List of header titles to write in the row.
    start_col : int
        The starting column index for the header.
    row : int
        The row index where the header will be placed.
    height : int, optional
        Row height for the header.
    width : int, optional
        Column width for each header cell.
    color : str or None, optional
        Background color for the header cells. Can be a key in `color_dict` or a hex color string.
        Default is `GREEN_CELL`.
    text_color : str or None, optional
        Font color for header text. Default is white ('D9FFFFFF').
    border_style : str, optional
        Border style to apply to each header cell (e.g., 'thin', 'medium').

    Notes:
    ------
    - All header cells are center-aligned and bold by default.
    - If `color` is in `color_dict`, the corresponding PatternFill is used.
    - This function is useful for creating consistent, styled headers in Excel reports.
    """
    for col_num, option in enumerate(data, start_col):
        cell = worksheet.cell(row=row, column=col_num, value=option)
        col_letter = get_column_letter(col_num)
        cell.alignment = Alignment_CELL
        if color is not None:
            if color in color_dict:
                cell.fill = color_dict[color]
            else:
                cell.fill = PatternFill(start_color=color, fill_type="solid")
        else:
            cell.fill = CREAM_CELL
        if text_color is not None:
            cell.font = Font(size=9, bold=True, color=text_color)
        else:
            cell.font = Font(size=9, bold=True, color='D9FFFFFF')
        if height is not None:
            worksheet.row_dimensions[row].height = height
        if width is not None:
            worksheet.column_dimensions[col_letter].width = width
        if border_style is not None:
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style=border_style),
                right=openpyxl.styles.Side(style=border_style),
                top=openpyxl.styles.Side(style=border_style),
                bottom=openpyxl.styles.Side(style=border_style)
            )


def create_header_freez(
        worksheet, data, start_col, row, header_row, height=None, width=None, len_with=None,
        different_cell=None, color=None
):
    """
    Create a styled header row in an Excel worksheet with freeze panes and auto-filter.

    Parameters:
    -----------
    worksheet : openpyxl.worksheet.worksheet.Worksheet
        The worksheet where the header will be created.
    data : list
        List of header titles to write in the row.
    start_col : int
        Starting column index for the header.
    row : int
        Row index where the header will be placed.
    header_row : int
        Row index to freeze (top row visible when scrolling).
    height : int, optional
        Row height for the header.
    width : int, optional
        Column width for each header cell.
    len_with : int, optional
        Additional width to add based on the length of the header text.
    different_cell : any, optional
        If a header value matches `different_cell`, its background will be red.
    color : str, optional
        Background color for the header cells. Can be a key in `color_dict` or a hex color string.
        Default is `GREEN_CELL`.

    Notes:
    ------
    - All header cells are center-aligned by default using `Alignment_CELL`.
    - Header row is frozen at `header_row` to keep it visible when scrolling.
    - Auto-filter is applied to the range from the first column to the last used column.
    - This function is useful for creating **Excel tables with fixed headers and filters**.
    """
    for col_num, option in enumerate(data, start_col):
        col_letter = get_column_letter(col_num)
        cell = worksheet.cell(row=row, column=col_num, value=option)
        cell.alignment = Alignment_CELL
        cell.fill = LIGHT_CREAM_CELL
        if color is not None:
            if color in color_dict:
                cell.fill = color_dict[color]
            else:
                cell.fill = PatternFill(start_color=color, fill_type="solid")

        if height is not None:
            worksheet.row_dimensions[row].height = height
            if len(option) > worksheet.column_dimensions[col_letter].width:
                worksheet.column_dimensions[col_letter].width = len(option) + 2
        if width is not None:
            worksheet.column_dimensions[col_letter].width = width
        if len_with is not None:
            if len(option) > worksheet.column_dimensions[col_letter].width:
                worksheet.column_dimensions[col_letter].width = len(option) + 3
        if different_cell is not None:
            if option == different_cell:
                cell.fill = PatternFill(start_color="C00000", fill_type="solid")
        worksheet.freeze_panes = worksheet[f'A{header_row}']
        max_col = worksheet.max_column
        range_str = f'A{header_row - 1}:{get_column_letter(max_col)}{worksheet.max_row}'
        worksheet.auto_filter.ref = range_str
