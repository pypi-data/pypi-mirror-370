import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from .styles import *


def create_value(worksheet, data, start_col, row, border_style=None, m=None, height=None, color=None, width=None,
                 different_cell=None, different_value=None, item_num=None, item_color=None, m_color=None):
    """
      Write a list of values into an Excel worksheet with optional formatting.

      Parameters:
      -----------
      worksheet : openpyxl.worksheet.worksheet.Worksheet
          The worksheet where values will be written.
      data : list
          List of values to write into the row.
      start_col : int
          Starting row index for writing values.
      row : int
          Starting column index for writing values.
      border_style : str, optional
          Border style to apply to each cell (e.g., 'thin', 'medium').
      m : int, optional
          Used to apply alternating row color; if `m % 2 != 0`, a light blue fill is applied.
      height : int, optional
          Row height.
      color : str, optional
          Cell background color (predefined colors like 'green', 'red', etc.).
      width : int, optional
          Column width.
      different_cell : int, optional
          Index of a cell in `data` to compare with `different_value`.
      different_value : any, optional
          If `data[different_cell] == different_value`, the cell is filled with red.
      item_num : int, optional
          Index of a specific item to apply `item_color`.
      item_color : PatternFill, optional
          Custom fill color for the item specified by `item_num`.

      Notes:
      ------
      - Numeric values are formatted with thousands separator ('#,###') if not zero.
      - Alignment is set to a predefined center alignment (`Alignment_CELL`).
      - Font size is set to 10 and bold by default.
      - `color_dict` is used for mapping color strings to actual fills.
      """

    for item in range(len(data)):
        cell = worksheet.cell(row=start_col, column=item + row, value=data[item])
        cell.alignment = Alignment_CELL

        if border_style:
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style=border_style),
                right=openpyxl.styles.Side(style=border_style),
                top=openpyxl.styles.Side(style=border_style),
                bottom=openpyxl.styles.Side(style=border_style)
            )

        value = data[item]
        if isinstance(value, (int, float)) and value != 0:
            cell.number_format = '#,###'
        else:
            cell.value = value

        cell.font = Font(size=10, bold=True)

        if m is not None and m % 2 == 0:
            if m_color:
                cell.fill = PatternFill(start_color=m_color, fill_type="solid")
            else:
                cell.fill = PatternFill(start_color=VERY_LIGHT_CREAM_CELL, fill_type="solid")

        if height is not None:
            worksheet.row_dimensions[start_col + 1].height = height

        if item_num is not None and item == item_num:
            if item_color:
                cell.fill = item_color
        elif color in color_dict:
            cell.fill = color_dict[color]

        if different_cell is not None and data[different_cell] == different_value:
            cell.fill = RED_CELL

        if width is not None:
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(item + row)].width = width
