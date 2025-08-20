"""Formula parsing and analysis utilities for spreadsheets."""

from typing import Dict, List, Set, Tuple, Optional, Any
import re
import pandas as pd


class FormulaParser:
    """
    Extracts, analyzes and simplifies Excel formulas from spreadsheets.
    
    This class provides utilities to:
    1. Extract formulas from Excel files
    2. Analyze cell dependencies based on formulas
    3. Generate simplified explanations of complex formulas
    4. Optimize formula encoding for LLMs
    """

    # Common Excel formula patterns
    SUM_PATTERN = re.compile(r'SUM\((.*?)\)')
    VLOOKUP_PATTERN = re.compile(r'VLOOKUP\((.*?)\)')
    CELL_REF_PATTERN = re.compile(r'([A-Z]+[0-9]+|[A-Z]+\:[A-Z]+|[0-9]+\:[0-9]+|[A-Z]+[0-9]+\:[A-Z]+[0-9]+)')
    
    def __init__(self):
        """Initialize the formula parser."""
        self.formula_map = {}  # Maps cell address to formula
        self.dependency_graph = {}  # Maps cell to its dependencies
        self.reverse_dependency = {}  # Maps cell to cells that depend on it
    
    def extract_formulas(self, excel_path: str) -> Dict[str, str]:
        """
        Extract all formulas from an Excel file.
        
        Args:
            excel_path: Path to the Excel file
            
        Returns:
            Dictionary mapping cell addresses to formulas
        """
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(excel_path, data_only=False)
            
            formulas = {}
            
            # Extract formulas from each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith('='):
                            # Store with sheet name for multi-sheet support
                            cell_address = f"{sheet_name}!{cell.coordinate}"
                            formulas[cell_address] = cell.value
            
            self.formula_map = formulas
            return formulas
            
        except ImportError:
            raise ImportError("openpyxl is required for formula extraction. Install with 'pip install openpyxl'")
    
    def build_dependency_graph(self) -> Dict[str, Set[str]]:
        """
        Build a graph of cell dependencies based on extracted formulas.
        
        Returns:
            Dictionary mapping cells to their dependencies
        """
        dependency_graph = {}
        reverse_dependency = {}
        
        for cell, formula in self.formula_map.items():
            # Find all cell references in the formula
            references = self.extract_cell_references(formula)
            
            # Add to dependency graph
            dependency_graph[cell] = set(references)
            
            # Build reverse dependency (which cells depend on this cell)
            for ref in references:
                if ref not in reverse_dependency:
                    reverse_dependency[ref] = set()
                reverse_dependency[ref].add(cell)
        
        self.dependency_graph = dependency_graph
        self.reverse_dependency = reverse_dependency
        return dependency_graph
    
    def extract_cell_references(self, formula: str) -> List[str]:
        """
        Extract all cell references from a formula.
        
        Args:
            formula: Excel formula string
            
        Returns:
            List of cell references found in the formula
        """
        # Remove the equals sign
        if formula.startswith('='):
            formula = formula[1:]
            
        # Find all cell references using regex
        references = self.CELL_REF_PATTERN.findall(formula)
        return references
    
    def simplify_formula(self, formula: str) -> str:
        """
        Generate a simplified explanation of a complex formula.
        
        Args:
            formula: Excel formula to simplify
            
        Returns:
            Human-readable explanation of the formula
        """
        if not formula.startswith('='):
            return f"Static value: {formula}"
            
        formula = formula[1:]  # Remove equals sign
        
        # Check for common patterns and provide explanations
        if "SUM" in formula:
            match = self.SUM_PATTERN.search(formula)
            if match:
                range_str = match.group(1)
                return f"Sum of values in range {range_str}"
                
        elif "AVERAGE" in formula:
            return f"Average of values in {formula.split('(')[1].split(')')[0]}"
            
        elif "VLOOKUP" in formula:
            match = self.VLOOKUP_PATTERN.search(formula)
            if match:
                args = match.group(1).split(',')
                if len(args) >= 3:
                    lookup_value = args[0].strip()
                    table_range = args[1].strip()
                    col_index = args[2].strip()
                    return f"Lookup '{lookup_value}' in table {table_range} and return value from column {col_index}"
        
        # Generic explanation for other formulas
        return f"Formula: {formula}"
    
    def get_formula_impact(self, cell_address: str) -> Dict[str, Any]:
        """
        Analyze the impact of a formula cell.
        
        Args:
            cell_address: Address of the cell to analyze
            
        Returns:
            Dictionary with impact analysis
        """
        if not self.dependency_graph or not self.reverse_dependency:
            self.build_dependency_graph()
            
        # Get direct dependencies (cells this formula uses)
        dependencies = self.dependency_graph.get(cell_address, set())
        
        # Get cells that depend on this cell (reverse dependencies)
        dependents = self.reverse_dependency.get(cell_address, set())
        
        # Get the formula
        formula = self.formula_map.get(cell_address, "")
        
        return {
            "cell": cell_address,
            "formula": formula,
            "simplified_explanation": self.simplify_formula(formula),
            "dependencies": list(dependencies),
            "dependents": list(dependents),
            "is_leaf": len(dependencies) == 0,
            "is_root": len(dependents) == 0,
            "dependency_depth": self._calculate_dependency_depth(cell_address)
        }
    
    def _calculate_dependency_depth(self, cell_address: str) -> int:
        """Calculate the maximum depth of the dependency chain."""
        if cell_address not in self.dependency_graph or not self.dependency_graph[cell_address]:
            return 0
            
        max_depth = 0
        for dep in self.dependency_graph[cell_address]:
            depth = 1 + self._calculate_dependency_depth(dep)
            max_depth = max(max_depth, depth)
            
        return max_depth
    
    def encode_formulas_for_llm(self, formulas: Dict[str, str] = None) -> str:
        """
        Generate LLM-friendly encoding of formulas.
        
        Args:
            formulas: Dictionary of cell formulas, or None to use stored formulas
            
        Returns:
            Formatted string representation of formulas for LLM consumption
        """
        if formulas is None:
            formulas = self.formula_map
            
        if not formulas:
            return "No formulas found"
            
        lines = ["## Spreadsheet Formulas"]
        
        # Group similar formulas
        formula_groups = {}
        for cell, formula in formulas.items():
            if formula not in formula_groups:
                formula_groups[formula] = []
            formula_groups[formula].append(cell)
        
        # Output grouped by formula type for efficiency
        for formula, cells in formula_groups.items():
            simplified = self.simplify_formula(formula)
            cell_list = ", ".join(cells[:5])
            if len(cells) > 5:
                cell_list += f" (+{len(cells)-5} more)"
            lines.append(f"- {simplified}")
            lines.append(f"  - Cells: {cell_list}")
            lines.append(f"  - Formula: `{formula}`")
            lines.append("")
            
        return "\n".join(lines)


class FormulaDependencyAnalyzer:
    """
    Specialized analyzer for formula dependencies and calculation chains.
    
    This class provides advanced analysis of formula relationships and
    calculation flows within spreadsheets.
    """
    
    def __init__(self, formula_parser: FormulaParser = None):
        """Initialize with an optional formula parser."""
        self.parser = formula_parser or FormulaParser()
        
    def find_calculation_chains(self) -> List[List[str]]:
        """
        Identify calculation chains (sequences of dependent formulas).
        
        Returns:
            List of calculation chains (each a list of cell addresses)
        """
        if not self.parser.dependency_graph:
            raise ValueError("Dependency graph not built. Call extract_formulas and build_dependency_graph first.")
            
        # Find root cells (cells that nothing depends on)
        root_cells = set()
        for cell in self.parser.reverse_dependency:
            if cell not in self.parser.reverse_dependency or not self.parser.reverse_dependency[cell]:
                root_cells.add(cell)
                
        # Build chains from each root
        chains = []
        for root in root_cells:
            chain = self._build_chain_from_cell(root)
            if chain:
                chains.append(chain)
                
        return chains
    
    def _build_chain_from_cell(self, start_cell: str) -> List[str]:
        """Build a calculation chain starting from a specific cell."""
        chain = [start_cell]
        current = start_cell
        
        # Follow the dependency chain
        while current in self.parser.dependency_graph and self.parser.dependency_graph[current]:
            # For simplicity, just follow first dependency
            # More complex chains would need tree traversal
            next_cell = next(iter(self.parser.dependency_graph[current]))
            if next_cell in chain:  # Avoid circular references
                break
            chain.append(next_cell)
            current = next_cell
            
        return chain
    
    def identify_critical_cells(self) -> List[str]:
        """
        Identify critical cells that many calculations depend on.
        
        Returns:
            List of cell addresses sorted by importance
        """
        if not self.parser.reverse_dependency:
            raise ValueError("Dependency analysis not performed")
            
        # Count how many cells depend on each cell
        importance = {}
        for cell, dependents in self.parser.reverse_dependency.items():
            importance[cell] = len(dependents)
            
        # Sort by importance (number of dependents)
        critical_cells = sorted(importance.items(), key=lambda x: x[1], reverse=True)
        return [cell for cell, count in critical_cells if count > 0]
