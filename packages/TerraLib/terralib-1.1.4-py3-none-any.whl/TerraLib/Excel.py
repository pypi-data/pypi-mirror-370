# coding=utf-8

from copy import copy
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.styles import Border, Side, PatternFill, Alignment
from enum import Enum
from pathlib import Path

MERGED_VALUE = "merged"


class Color(Enum):
    BLACK = COLOR_INDEX[0]
    WHITE = COLOR_INDEX[1]
    RED = COLOR_INDEX[2]
    GREEN = COLOR_INDEX[3]
    BLUE = COLOR_INDEX[4]
    YELLOW = COLOR_INDEX[5]
    PINK = COLOR_INDEX[6]
    LIGHT_BLUE = COLOR_INDEX[7]
    GRAY = COLOR_INDEX[22]
    ORANGE = COLOR_INDEX[52]


class Fill(Enum):
    BLACK = PatternFill("solid", fgColor=Color.BLACK.value)
    WHITE = PatternFill("solid", fgColor=Color.WHITE.value)
    RED = PatternFill("solid", fgColor=Color.RED.value)
    GREEN = PatternFill("solid", fgColor=Color.GREEN.value)
    BLUE = PatternFill("solid", fgColor=Color.BLUE.value)
    YELLOW = PatternFill("solid", fgColor=Color.YELLOW.value)
    PINK = PatternFill("solid", fgColor=Color.PINK.value)
    LIGHT_BLUE = PatternFill("solid", fgColor=Color.LIGHT_BLUE.value)
    GRAY = PatternFill("solid", fgColor=Color.GRAY.value)
    ORANGE = PatternFill("solid", fgColor=Color.ORANGE.value)


class HAlign(Enum):
    LEFT = "left"
    RIGHT = "right"
    CENTER = "center"
    JUSTIFY = "justify"


class VAlign(Enum):
    TOP = "top"
    BOTTOM = "bottom"
    CENTER = "center"
    JUSTIFY = "justify"


default_border = Border(left=Side(style='thin', color=Color.BLACK.value),
                        right=Side(style='thin', color=Color.BLACK.value),
                        top=Side(style='thin', color=Color.BLACK.value),
                        bottom=Side(style='thin', color=Color.BLACK.value),
                        vertical=Side(style='thin', color=Color.BLACK.value),
                        horizontal=Side(style='thin', color=Color.BLACK.value))


class ExcelType(Enum):
    XLS = "xls"
    XLSX = "xlsx"
    UNKNOWN = "unknown"


class MyBook:
    def __init__(self, path=None):
        """
        Load or create a workbook, if path is None will create a new xlsx else will load the excel file
        :param path:
        """
        self.pre = None
        self.path = Path(path) if path is not None else None

        if self.path is not None and not self.path.exists():
            raise IOError(f"Not exists {str(self.path)}!")

        if self.path is None:
            # create a new xlsx, only support xlsx
            self.excel_type = ExcelType.XLSX
            from openpyxl import Workbook
            self.pre = Workbook()
        else:
            # load workbook , support xlsx and xls
            if self.path.suffix in [".xlsx", ".xlsm"]:
                self.excel_type = ExcelType.XLSX
                from openpyxl import load_workbook
                self.pre = load_workbook(self.path, data_only=True)
            elif self.path.suffix == ".xls":
                self.excel_type = ExcelType.XLS
                from xlrd import open_workbook
                self.pre = open_workbook(self.path, formatting_info=True)
            else:
                raise TypeError(f"Not support file type:{self.path.suffix}")

    def sheet_by_index(self, index):
        if self.excel_type == ExcelType.XLS:
            ws = self.pre.sheet_by_index(index)
        else:
            ws = self.pre.worksheets[index]
        return MySheet(ws, self)

    def sheet_by_name(self, name):
        if self.excel_type == ExcelType.XLS:
            ws = self.pre.sheet_by_name(name)
        else:
            ws = self.pre.get_sheet_by_name(name)
        return MySheet(ws, self)

    def sheet_names(self):
        if self.excel_type == ExcelType.XLS:
            return self.pre.sheet_names()
        else:
            return self.pre.get_sheet_names()

    def sheets(self):
        if self.excel_type == ExcelType.XLS:
            ws = self.pre.sheets()
        else:
            ws = self.pre.worksheets
        return (MySheet(x, self) for x in ws)

    def add_sheet(self, name):
        if self.excel_type == ExcelType.XLS:
            raise (f"Not support {ExcelType.XLS.value}.")
        else:
            ws = self.pre.create_sheet(title=name)
        return MySheet(ws, self)

    def remove_sheet(self, name):
        if self.excel_type == ExcelType.XLS:
            raise (f"Not support {ExcelType.XLS.value}.")
        else:
            ws = self.sheet_by_name(name).pre
            self.pre.remove(ws)
        return None

    def save(self, new_file=None):
        if new_file:
            save_path = new_file
        elif self.path:
            save_path = self.path
        else:
            save_path = "temp.xlsx"
        self.pre.save(save_path)


class MySheet:
    def __init__(self, ws, parent_wb):
        self.wb = parent_wb
        self.pre = ws
        self.excel_type = self.wb.excel_type
        self.name = ws.name if self.excel_type == ExcelType.XLS else ws.title
        self.nrows = ws.nrows if self.excel_type == ExcelType.XLS else ws.max_row
        self.ncols = ws.ncols if self.excel_type == ExcelType.XLS else ws.max_column

    def get_cell(self, row, col):
        if self.excel_type == ExcelType.XLS:
            return MyCell(self.pre.cell(row, col), self, row, col)
        else:
            return MyCell(self.pre.cell(row=row + 1, column=col + 1), self, row, col)

    def gen_cells(self, min_idx=0, max_idx=None, start=0, end=None, by_col=False, flatten=True):
        """
        Get multi cells
        :param min_idx:
        :param max_idx:
        :param start:
        :param end:
        :param by_col:
        :param flatten: True, yield cell one by one otherwise yield an generator with cells
        :return:
        """
        if not by_col:
            act_max = max_idx if max_idx or max_idx == 0 else self.nrows - 1
            act_end = end if end or end == 0 else self.ncols - 1
        else:
            act_max = max_idx if max_idx or max_idx == 0 else self.ncols - 1
            act_end = end if end or end == 0 else self.nrows - 1
        idx_list = range(min_idx, act_max + 1)
        if self.excel_type == ExcelType.XLS:
            if not by_col:
                for act_idx in idx_list:
                    one_slice = (MyCell(x, self, act_idx, start + y) for y, x in
                                 enumerate(self.pre.row_slice(act_idx, start, act_end + 1)))
                    if not flatten:
                        yield one_slice
                    else:
                        yield from one_slice
            else:
                for act_idx in idx_list:
                    one_slice = (MyCell(x, self, start + y, act_idx) for y, x in
                                 enumerate(self.pre.col_slice(act_idx, start, act_end + 1)))
                    if not flatten:
                        yield one_slice
                    else:
                        yield from one_slice
        else:
            if not by_col:
                cells = self.pre.iter_rows(min_row=idx_list[0] + 1, max_row=idx_list[-1] + 1, min_col=start + 1,
                                           max_col=act_end + 1)
            else:
                cells = self.pre.iter_cols(min_row=start + 1, max_row=act_end + 1, min_col=idx_list[0] + 1,
                                           max_col=idx_list[-1] + 1)
            for one_row_col in cells:  # generate one row/column
                one_slice = (MyCell(x, self) for x in one_row_col)
                if not flatten:
                    yield one_slice
                else:
                    yield from one_slice

    def set_value(self, data, start_idx=(0, 0), colour=None, fill=None, border=False, halign=HAlign.JUSTIFY,
                  valign=VAlign.JUSTIFY):
        """
        set values for cells
        :param data:
        :param start_idx:
        :param colour:
        :param fill:
        :param border:
        :param halign:
        :param valign:
        :return:
        """
        ''' if by_col = False , first dimi write to one row '''
        for i, one_row in enumerate(data):
            if isinstance(one_row, list) or isinstance(one_row, tuple):
                for j, one in enumerate(one_row):
                    cell_obj = self.pre.cell(row=i + 1 + start_idx[0], column=j + 1 + start_idx[1])
                    cell_obj.value = one
                    cell_obj.alignment = Alignment(horizontal=halign.value, vertical=valign.value)
                    if colour:
                        tmp_font = copy(cell_obj.font)
                        tmp_font.color = colour.value if isinstance(colour, Color) else colour
                        cell_obj.font = tmp_font
                    if fill:
                        cell_obj.fill = fill.value
                    if border:
                        cell_obj.border = default_border

            else:
                cell_obj = self.pre.cell(row=i + 1 + start_idx[0], column=1 + start_idx[1])
                cell_obj.value = one_row
                cell_obj.alignment = Alignment(horizontal=halign.value, vertical=valign.value)
                if colour:
                    tmp_font = copy(cell_obj.font)
                    tmp_font.color = colour.value if isinstance(colour, Color) else colour
                    cell_obj.font = tmp_font
                if fill:
                    cell_obj.fill = fill.value
                if border:
                    cell_obj.border = default_border

    def set_style_auto_width(self, cols=None):
        """
        set column width
        :param cols: specify cols like [0,1,2] , None means all columns
        :return:
        """
        for i, one_col in enumerate(self.pre.columns):
            col_letter = get_column_letter(i + 1)
            if cols and i not in cols:
                continue
            max_col_width = max([len(str(one_col[x].value)) for x in range(len(one_col))])
            self.pre.column_dimensions[col_letter].width = 100 if max_col_width > 100 else max_col_width + 4

    def find(self, text, start_row=0, start_col=0, whole_match=True, by_row=True, ignore_case=False, silent=False):
        """
        add by terra,used to find a cell which match the text (whole or part)
        :param text: the text need to find
        :param start_row:
        :param start_col:
        :param whole_match: whole match or part match ,default is whole match
        :param by_row: by_row or by_col to search,default is by row
        :param ignore_case: ignore case or not,default is not ignore case
        :param silent: if True, will not print warning message when not found
        :return: if found return a cell else None
        """
        if by_row:
            for rowIndex in range(start_row, self.nrows):
                for colIndex in range(start_col, self.ncols):
                    cell = self.get_cell(rowIndex, colIndex)
                    cell_val = cell.value
                    if ignore_case:
                        cell_val = cell_val.lower()
                        text = text.lower()
                    if whole_match:
                        if cell_val == text:
                            return cell
                    else:
                        if text in cell_val:
                            return cell
        else:
            for colIndex in range(start_col, self.ncols):
                for rowIndex in range(start_row, self.nrows):
                    cell = self.get_cell(rowIndex, colIndex)
                    cell_val = cell.value
                    if ignore_case:
                        cell_val = cell_val.lower()
                        text = text.lower()
                    if whole_match:
                        if cell_val == text:
                            return cell
                    else:
                        if text in cell_val:
                            return cell
        if not silent:
            print("Warning: Can not find [%s] in sheet %s！" % (text, self.name))
        return None

    def merge_cells(self, cell1, cell2, border=True):
        """
        merge cells, for openpyxl only
        :param cell1:
        :param cell2:
        :param border:
        :return:
        """
        if border:
            cell1.pre.border = default_border
            cell2.pre.border = default_border
        self.pre.merge_cells(f"{cell1.addr}:{cell2.addr}")

    def group_rows(self, row_start, row_end, outline_level=1, hidden=False):
        """
        group rows, for openpyxl only
        :param row_start: the start row number, this row will be grouped
        :param row_end:  the last row number, this row will be grouped
        :param outline_level: group outline level,1-7,if set multi level, must set larger number first
        :param hidden: whether hiddern the group
        :return:
        """
        self.pre.row_dimensions.group(row_start + 1, row_end + 1, outline_level=outline_level, hidden=hidden)


class MyCell:
    def __init__(self, cell, parent_ws, row_idx=0, col_idx=0):
        self.pre = cell
        self.ws = parent_ws
        self.merge_val = None
        self.merge_range = None
        self.excel_type = self.ws.excel_type
        if self.excel_type == ExcelType.XLS:
            self.row_idx, self.col_idx = row_idx, col_idx
            self.trans_from_xlrd(cell)
        else:
            self.row_idx, self.col_idx = cell.row - 1, cell.column - 1
            self.trans_from_pyxl(cell)

    def __repr__(self):
        return '{0.addr} : {0.value} {0.bcolour}'.format(self)

    @property
    def comment(self):
        if self.excel_type == ExcelType.XLSX:
            if hasattr(self.pre, "comment"):
                return self.pre.comment
            else:
                return None
        else:
            if hasattr(self.ws.pre, "cell_note_map"):
                return self.ws.pre.cell_note_map.get((self.row_idx, self.col_idx))
            else:
                return None

    def trans_from_xlrd(self, cell):
        if cell.ctype in (2, 3) and int(cell.value) == cell.value:
            cell.value = int(cell.value)
        self.value = str(cell.value).strip()
        self.ctype = self.xlrd_ctype(cell.ctype)
        cell_fmt = self.ws.wb.pre.xf_list[cell.xf_index]
        font_fmt = self.ws.wb.pre.font_list[cell_fmt.font_index]
        self.bcolour = cell_fmt.background.pattern_colour_index
        self.bcolour_rgb = self.ws.wb.pre.colour_map[self.bcolour]
        self.bcolour_hex = self.__rgb_to_hex_color(self.bcolour_rgb)
        self.fcolour = font_fmt.colour_index
        self.fcolour_rgb = self.ws.wb.pre.colour_map[self.fcolour]
        self.fcolour_hex = self.__rgb_to_hex_color(self.fcolour_rgb)
        self.strike = font_fmt.struck_out
        self.merged = False
        for (row, row_range, col, col_range) in self.ws.pre.merged_cells:
            if self.row_idx >= row and self.row_idx < row_range and self.col_idx >= col and self.col_idx < col_range:
                if self.row_idx == row and self.col_idx == col:
                    break  # change continue to break by terra 2020/10/20
                self.merged = True
                base_cell = self.ws.pre.cell(row, col)
                self.merge_val = str(base_cell.value).strip() if base_cell.value is not None else ""
                self.merge_range = (row, row_range - 1, col, col_range - 1)
                self.value = MERGED_VALUE  # speicial nam for parse
                break  # change continue to break by terra 2020/10/20
        from xlrd import cellname
        self.addr = cellname(self.row_idx, self.col_idx)

    def trans_from_pyxl(self, cell):
        self.value = str(cell.value).strip() if cell.value is not None else ""
        self.ctype = 'text'  ## default always using text in openpyxl , update later
        self.bcolour = cell.fill.bgColor.value
        self.fcolour = cell.font.color
        self.strike = cell.font.strike
        self.addr = cell.coordinate
        self.merged = False  # add by terra 2020/9/23 openpyxl
        for cell_range in self.ws.pre.merged_cells:
            if cell_range.min_row <= self.row_idx + 1 <= cell_range.max_row and cell_range.min_col <= self.col_idx + 1 <= cell_range.max_col:
                if self.row_idx + 1 == cell_range.min_row and self.col_idx + 1 == cell_range.min_col:
                    break  # change continue to break by terra 2020/10/20
                self.merged = True
                base_cell = self.ws.pre.cell(cell_range.min_row, cell_range.min_col)
                self.merge_val = str(base_cell.value).strip() if base_cell.value is not None else ""
                self.merge_range = (
                    cell_range.min_row - 1, cell_range.max_row - 1, cell_range.min_col - 1, cell_range.max_col - 1)
                self.value = MERGED_VALUE
                break  # change continue to break by terra 2020/10/20

    def xlrd_ctype(self, ctype):
        return 'text'

    def set_value(self, value, colour=None):
        if self.excel_type == ExcelType.XLSX:
            self.pre.value = value
            if colour:
                self.set_style_fcolour(colour)
        else:
            raise NotImplemented("Not support for xlrd.")

    def set_style_bcolour(self, fill):
        if self.excel_type == ExcelType.XLSX:
            self.pre.fill = fill.value
        else:
            raise NotImplemented("Not support for xlrd.")

    def set_style_fcolour(self, colour):
        if self.excel_type == ExcelType.XLSX:
            tmp_font = copy(self.pre.font)
            tmp_font.color = colour
            self.pre.font = tmp_font
        else:
            raise NotImplemented("Not support for xlrd.")

    def set_style_alignment(self, halign=HAlign.JUSTIFY, valign=VAlign.JUSTIFY):
        if self.excel_type == ExcelType.XLSX:
            self.pre.alignment = Alignment(horizontal=halign.value, vertical=valign.value)
        else:
            raise NotImplemented("Not support for xlrd.")

    def set_style_border(self, pos=None, style='thin'):
        """
        :param pos: top, left, right, bottom , None means all
        :param style:
        :return:
        """
        if self.excel_type == ExcelType.XLSX:
            from openpyxl.styles import Side
            new_side = Side(border_style=style, color=Color.BLACK.value)
            tmp_border = copy(self.pre.border)
            if not pos:
                pos = ['top', 'left', 'right', 'bottom']
            for one in pos:
                assert one in ('top', 'left', 'right', 'bottom'), "border position wrong : {}".format(one)
                setattr(tmp_border, one, new_side)
            self.pre.border = tmp_border
        else:
            raise NotImplemented("Not support for xlrd.")

    def __rgb_to_hex_color(self, rgb):
        if rgb:
            r, g, b = rgb
            return f"#{'{:02X}'.format(r)}{'{:02X}'.format(g)}{'{:02X}'.format(b)}"
        else:
            return None

    def __hex_color_to_rgb(self, hex_color):

        hex_color = hex_color.strip("#")
        if len(hex_color) == 8:
            hex_color = hex_color[2:]
        return int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)


if __name__ == '__main__':
    wb = MyBook(r"C:\Users\terra\Desktop\test.xlsx")
    ws = wb.sheet_by_index(0)
    for one in ws.gen_cells():
        print(f"{one.addr} {one.value} {one.merge_val} {one.merge_range}")
