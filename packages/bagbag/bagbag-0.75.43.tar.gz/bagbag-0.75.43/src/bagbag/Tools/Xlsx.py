import openpyxl
import os

#print("load " + '/'.join(__file__.split('/')[-2:]))

class Reader():
    def __init__(self, fpath:str, withHeader:bool=True):
        self.fpath = fpath 
        self.wb = openpyxl.load_workbook(filename=fpath)
        self.ws = self.wb.active
        self.iws = self.ws.iter_rows()

        self.withHeader = withHeader
        self.headers = None
        if self.withHeader:
            self.headers = [j for j in filter(lambda x: x != None, [i.value for i in next(self.iws)])]

    def SetHeaders(self, *headers):
        self.headers = headers
    
    def Read(self) -> dict:
        r = [j for j in filter(lambda x: x != None, [i.value for i in next(self.iws)])]

        if len(r) == 0:
            raise StopIteration

        row = {}
        for idx in range(len(self.headers)):
            try:
                row[self.headers[idx]] = r[idx]
            except IndexError:
                self.Close()
                row[self.headers[idx]] = "" 
        
        return row
    
    def __iter__(self):
        while True:
            try:
                yield self.Read()
            except StopIteration:
                return 
    
    def Close(self):
        self.wb.close()

class Writer():
    def __init__(self, fpath:str, mode:str="w"):
        self.fpath = fpath
        self.fdmode = mode
        self.headers = None
        if mode == "a" and not os.path.exists(fpath):
            self.fdmode = "w"
            
        if self.fdmode != "w":
            try:
                r = Reader(fpath)
                self.headers = r.headers
                r.Close()
            except StopIteration:
                self.fdmode = "w"
            self.wb = openpyxl.load_workbook(filename=fpath)
        else:
            self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def SetHeaders(self, *headers):
        self.headers = headers
        if self.fdmode == "w":
            self.ws.append(self.headers)
    
    def Write(self, row:dict[str]):
        r = []
        for header in self.headers:
            if header in row:
                r.append(row[header])
            else:
                r.append("")
        
        self.ws.append(r)

    def Close(self):
        self.wb.save(self.fpath)
        self.wb.close()
    
    def Flush(self):
        self.wb.save(self.fpath)
    
    def __enter__(self):
        return self 
    
    def __exit__(self, exc_type, exc_value, traceback):
        try:
            self.Close()
        except:
            pass

# class Xlsx:
#     Reader
#     Writer

if __name__ == "__main__":
    w = Writer("test.xlsx")

    w.SetHeaders("h1", "h2")

    w.Write({"h1": "v1", "h2": '"v2,kkk|'})
    w.Write({"h1": "v,1", "h2": '"v222'})
    w.Write({"h1": "3", "h2": '"99kkk'})

    w.Close()

    # test.csv
    # h1,h2
    # v1,"\"v2,kkk|"
    # "v,1",\"v222
    # 3,\"99kkk

    r = Reader("test.xlsx")
    print(r.Read()) # {'h1': 'v1', 'h2': '"v2,kkk|'}

    for row in r:
        print(row) 
        # {'h1': 'v,1', 'h2': '"v222'}
        # {'h1': '3', 'h2': '"99kkk'}
    
    w = Writer("test.xlsx", "a")
    w.Write({"h1": "4", "h2": '5'}) 
    w.Write({"h1": "6", "h3": '7'}) # 6,
    w.Close() # 保存

    w = Writer("test1.xlsx", "a")
    w.SetHeaders("h1", "h2")
    w.Write({"h1": "4", "h2": '5'}) 
    w.Write({"h1": "6", "h3": '7'}) # 6,
    w.Flush() # 保存
    w.Close() # 保存, 关闭